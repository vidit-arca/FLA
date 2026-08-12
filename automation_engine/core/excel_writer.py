import openpyxl
from openpyxl.utils import get_column_letter
import os

class ExcelWriter:
    def __init__(self, skeletal_path, output_path=None):
        self.skeletal_path = skeletal_path
        if output_path is None:
            # Save in the same folder as skeletal but with a new name
            dir_name = os.path.dirname(skeletal_path)
            self.output_path = os.path.join(dir_name, "FLA Return Populated.xlsx")
        else:
            self.output_path = output_path
            
    def write_values(self, cell_values):
        """Copies skeletal Excel to output_path, writes calculated/extracted values on the copy, and saves it."""
        import shutil
        
        if not os.path.exists(self.skeletal_path):
            raise FileNotFoundError(f"Skeletal Excel not found at {self.skeletal_path}")
            
        # Ensure output directory exists
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        
        # Copy skeletal template to the output path first
        print(f"[*] Copying skeletal template to: {self.output_path}")
        shutil.copy(self.skeletal_path, self.output_path)
        
        print(f"[*] Loading copied target Excel: {self.output_path}")
        wb = openpyxl.load_workbook(self.output_path, data_only=False)
        
        # Iterate over sections
        # Iterate over sections
        for section, cells in cell_values.items():
            if section not in wb.sheetnames:
                print(f"[!] Warning: Sheet '{section}' not found in skeletal Excel. Skipping...")
                continue
                
            sheet = wb[section]
            print(f"[*] Populating sheet '{section}' with {len(cells)} active fields...")
            
            # Specific coordinates that are merged in the skeletal template but need independent PY/FY values
            unmerge_targets = {
                "Section II": ["C5:G5", "C6:G6", "C11:G11", "C24:G24", "C30:G30", "C34:G34"],
                "Section III": ["C20:E20", "C23:E23", "C44:E44", "C47:E47"],
                "Section IV": ["C39:E39", "C42:E42"]
            }
            
            # Custom dynamic row insertion for Section III multiple investors and countries
            if section == "Section III":
                cells = self._process_section_3_dynamic(sheet, cells, unmerge_targets)
            
            if section in unmerge_targets:
                for rng_str in unmerge_targets[section]:
                    try:
                        # Check if it is currently merged in the sheet
                        for rng in list(sheet.merged_cells.ranges):
                            if rng.coord == rng_str:
                                sheet.unmerge_cells(rng_str)
                                # Set the top-left cell explicitly
                                top_left = rng_str.split(":")[0]
                                sheet[top_left] = "Auto-calculated"
                                break
                    except Exception as e:
                        print(f"[!] Warning: Could not unmerge {rng_str} in sheet {section}: {e}")
            
            # Map merged cells to their top-left parent cells
            merged_map = {}
            for rng in sheet.merged_cells.ranges:
                top_left = rng.start_cell.coordinate
                for r in range(rng.min_row, rng.max_row + 1):
                    for c in range(rng.min_col, rng.max_col + 1):
                        cell_coord = get_column_letter(c) + str(r)
                        merged_map[cell_coord] = top_left
            
            for coord, val in cells.items():
                resolved_coord = merged_map.get(coord, coord)
                try:
                    # Preserve template structure: do not overwrite cells containing "Auto-calculated"
                    existing_val = sheet[resolved_coord].value
                    if isinstance(existing_val, str) and "auto-calculated" in existing_val.lower():
                        # Skipping to retain the merged cell structure and visual "Auto-calculated" label
                        continue
                        
                    # If this is a percentage coordinate, convert to decimal and apply percentage formatting
                    if (section == "Section II" and resolved_coord in ["F24", "G24"]) or \
                       (section == "Section IV" and resolved_coord in ["E19", "F19"]):
                        try:
                            val_float = float(val)
                            val = val_float / 100.0
                            sheet[resolved_coord]
                        except Exception:
                            pass
                            
                    # Write to the resolved coordinate (always the top-left of the merged block)
                    sheet[resolved_coord] = val
                    
                    # Ensure font text color is Black (000000) so white theme text is not invisible!
                    from openpyxl.styles import Font
                    curr_font = sheet[resolved_coord].font
                    if curr_font:
                        sheet[resolved_coord].font = Font(
                            name=curr_font.name or "Aptos Narrow",
                            size=curr_font.size or 11,
                            bold=curr_font.bold,
                            italic=curr_font.italic,
                            color="000000"
                        )
                    else:
                        sheet[resolved_coord].font = Font(color="000000")
                    
                except Exception as e:
                    print(f"[!] Error writing {val} to cell {coord} (resolved: {resolved_coord}) in sheet {section}: {e}")



                    
        # Clear intermediate logic columns K, L, M for the client-facing compliance sheet
        if "compliance for Private " in wb.sheetnames:
            comp_sheet = wb["compliance for Private "]
            for row in range(1, 150):
                try: comp_sheet.cell(row=row, column=11).value = None # K
                except AttributeError: pass
                try: comp_sheet.cell(row=row, column=12).value = None # L
                except AttributeError: pass
                try: comp_sheet.cell(row=row, column=13).value = None # M
                except AttributeError: pass

        # Apply premium formatting fixes to prevent clipping, overlapping, or jagged borders
        self.beautify_layout(wb)
                    
        # Save output
        print(f"[+] Saving populated Excel to: {self.output_path}")
        wb.save(self.output_path)
        return self.output_path

    def beautify_layout(self, wb):
        """Applies premium formatting fixes to prevent clipping, overlapping, or jagged borders."""
        from openpyxl.styles import Border, Side, Alignment
        
        # 1. Section III Beautifications
        if "Section III" in wb.sheetnames:
            sheet = wb["Section III"]
            # Row 16 contains a very long description in C16. Increase row height to avoid clipping!
            sheet.row_dimensions[16].height = 48
            # Row 17 (data row) can have a normal, clean height
            sheet.row_dimensions[17].height = 24
            
            # Align filled cells beautifully
            for r in [17, 21, 22, 24, 25, 26]:
                for c in [2, 3, 4, 5, 6, 7]:
                    cell = sheet.cell(row=r, column=c)
                    # Align: left/center for text, center for country/numbers
                    if c in [2, 3]:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    elif c in [4, 5]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Unify table borders up to column G (columns 1 to 7) for rows 15 to 26
            # to remove the jaggedness on the right side of the sheet!
            for r in range(15, 27):
                for c in range(1, 8):
                    cell = sheet.cell(row=r, column=c)
                    current_border = cell.border
                    new_left = current_border.left if current_border.left.style else Side(style='thin', color='000000')
                    new_right = current_border.right if current_border.right.style else Side(style='thin', color='000000')
                    new_top = current_border.top if current_border.top.style else Side(style='thin', color='000000')
                    new_bottom = current_border.bottom if current_border.bottom.style else Side(style='thin', color='000000')
                    
                    cell.border = Border(left=new_left, right=new_right, top=new_top, bottom=new_bottom)
                    
        # 2. Section IV Beautifications
        if "Section IV" in wb.sheetnames:
            sheet = wb["Section IV"]
            # Set DIE 1 details row heights and alignment
            sheet.row_dimensions[19].height = 24
            
            # Align filled cells beautifully
            for r in [19, 23, 24, 26, 27, 28]:
                for c in [1, 2, 3, 4, 5, 6]:
                    cell = sheet.cell(row=r, column=c)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
            # Unify borders up to column G for rows 17 to 29
            for r in range(17, 30):
                for c in range(1, 8):
                    cell = sheet.cell(row=r, column=c)
                    current_border = cell.border
                    new_left = current_border.left if current_border.left.style else Side(style='thin', color='000000')
                    new_right = current_border.right if current_border.right.style else Side(style='thin', color='000000')
                    new_top = current_border.top if current_border.top.style else Side(style='thin', color='000000')
                    new_bottom = current_border.bottom if current_border.bottom.style else Side(style='thin', color='000000')
                    
                    cell.border = Border(left=new_left, right=new_right, top=new_top, bottom=new_bottom)

    def _process_section_3_dynamic(self, sheet, cells, unmerge_targets):
        import json
        import re
        import copy
        
        new_cells = {}
        fdi_json = cells.pop("fdi_investors_json", "[]")
        di_json = cells.pop("di_countries_json", "[]")
        
        try:
            fdi_investors = json.loads(fdi_json)
            di_countries = json.loads(di_json)
        except:
            fdi_investors = []
            di_countries = []
            
        # We need to process FDI blocks first, then DI blocks.
        fdi_blocks_to_add = max(0, len(fdi_investors) - 1)
        fdi_rows_per_block = 13
        fdi_shift = fdi_blocks_to_add * fdi_rows_per_block
        
        di_blocks_to_add = max(0, len(di_countries) - 1)
        di_rows_per_block = 12
        di_shift = di_blocks_to_add * di_rows_per_block
        
        merged_ranges = list(sheet.merged_cells.ranges)
        sheet.merged_cells.ranges = []
        
        # 1. Insert rows for FDI
        if fdi_shift > 0:
            print(f"[*] Section III: Inserting {fdi_shift} rows for {fdi_blocks_to_add} additional FDI investors...")
            sheet.insert_rows(28, fdi_shift)
            
        # 2. Insert rows for DI
        di_start_row = 39 + fdi_shift
        if di_shift > 0:
            print(f"[*] Section III: Inserting {di_shift} rows for {di_blocks_to_add} additional DI countries...")
            sheet.insert_rows(di_start_row + 12, di_shift)
            
        # 3. Handle merged cells shifting and duplication
        for r in merged_ranges:
            if r.min_row > 50:
                # Shift by both
                r.shift(row_shift=fdi_shift + di_shift)
                sheet.merged_cells.add(r)
            elif 39 <= r.min_row <= 50:
                # Shift down by FDI shift to get to new DI start
                r.shift(row_shift=fdi_shift)
                sheet.merged_cells.add(r)
                # Duplicate for DI
                for i in range(1, di_blocks_to_add + 1):
                    new_r = copy.copy(r)
                    new_r.shift(row_shift=i * di_rows_per_block)
                    sheet.merged_cells.add(new_r)
            elif 28 <= r.min_row <= 38:
                # Interstitial rows between FDI and DI (Rows 28 to 38)
                r.shift(row_shift=fdi_shift)
                sheet.merged_cells.add(r)
            elif 15 <= r.min_row <= 27:
                # FDI block
                sheet.merged_cells.add(r)
                # Duplicate for FDI
                for i in range(1, fdi_blocks_to_add + 1):
                    new_r = copy.copy(r)
                    new_r.shift(row_shift=i * fdi_rows_per_block)
                    sheet.merged_cells.add(new_r)
            else:
                sheet.merged_cells.add(r)
                
        # 4. Copy styles/borders for FDI blocks
        for i in range(1, fdi_blocks_to_add + 1):
            offset = i * fdi_rows_per_block
            for row_offset in range(fdi_rows_per_block):
                src_row = 15 + row_offset
                target_row = src_row + offset
                sheet.row_dimensions[target_row].height = sheet.row_dimensions[src_row].height
                for col_idx in range(1, sheet.max_column + 1):
                    src_cell = sheet.cell(row=src_row, column=col_idx)
                    dst_cell = sheet.cell(row=target_row, column=col_idx)
                    if type(dst_cell).__name__ == 'MergedCell':
                        continue
                    try:
                        dst_cell.value = src_cell.value
                    except Exception:
                        pass
                    if src_cell.has_style:
                        if src_cell.font: dst_cell.font = copy.copy(src_cell.font)
                        if src_cell.border: dst_cell.border = copy.copy(src_cell.border)
                        if src_cell.fill: dst_cell.fill = copy.copy(src_cell.fill)
                        if src_cell.alignment: dst_cell.alignment = copy.copy(src_cell.alignment)
                        dst_cell.number_format = src_cell.number_format
                        
        # 5. Copy styles/borders for DI blocks
        for i in range(1, di_blocks_to_add + 1):
            offset = i * di_rows_per_block
            for row_offset in range(di_rows_per_block):
                src_row = 39 + fdi_shift + row_offset
                target_row = src_row + offset
                sheet.row_dimensions[target_row].height = sheet.row_dimensions[src_row].height
                for col_idx in range(1, sheet.max_column + 1):
                    src_cell = sheet.cell(row=src_row, column=col_idx)
                    dst_cell = sheet.cell(row=target_row, column=col_idx)
                    if type(dst_cell).__name__ == 'MergedCell':
                        continue
                    try:
                        dst_cell.value = src_cell.value
                    except Exception:
                        pass
                    if src_cell.has_style:
                        if src_cell.font: dst_cell.font = copy.copy(src_cell.font)
                        if src_cell.border: dst_cell.border = copy.copy(src_cell.border)
                        if src_cell.fill: dst_cell.fill = copy.copy(src_cell.fill)
                        if src_cell.alignment: dst_cell.alignment = copy.copy(src_cell.alignment)
                        dst_cell.number_format = src_cell.number_format

        if not hasattr(self, 'sec3_pct_cells'):
            self.sec3_pct_cells = []
            
        # 6. Populate FDI blocks dynamically (Rows 15 to 27)
        for idx, inv in enumerate(fdi_investors):
            offset = idx * fdi_rows_per_block
            base = 15 + offset
            
            new_cells[f"B{base+2}"] = inv.get("name", "") # B17
            new_cells[f"C{base+2}"] = inv.get("country", "") # C17
            new_cells[f"D{base+2}"] = inv.get("equity_percent_py", 0) # D17
            new_cells[f"E{base+2}"] = inv.get("equity_percent_fy", 0) # E17
            self.sec3_pct_cells.extend([f"D{base+2}", f"E{base+2}"])
            
            new_cells[f"D{base+5}"] = inv.get("equity_capital_py", 0) # D20
            new_cells[f"E{base+5}"] = inv.get("equity_capital_fy", 0) # E20
            
            new_cells[f"D{base+6}"] = inv.get("liabilities_py", 0) # D21
            new_cells[f"E{base+6}"] = inv.get("liabilities_fy", 0) # E21
            
            new_cells[f"D{base+7}"] = inv.get("claims_py", 0) # D22
            new_cells[f"E{base+7}"] = inv.get("claims_fy", 0) # E22
            
            new_cells[f"D{base+8}"] = inv.get("other_capital_py", 0) # D23
            new_cells[f"E{base+8}"] = inv.get("other_capital_fy", 0) # E23
            
            new_cells[f"D{base+9}"] = inv.get("fallback_liabilities_py", 0) # D24
            new_cells[f"E{base+9}"] = inv.get("fallback_liabilities_fy", 0) # E24
            
            new_cells[f"D{base+10}"] = inv.get("fallback_claims_py", 0) # D25
            new_cells[f"E{base+10}"] = inv.get("fallback_claims_fy", 0) # E25
            
            new_cells[f"D{base+11}"] = 0.0 # D26 Disinvestments
            new_cells[f"E{base+11}"] = 0.0 # E26 Disinvestments
            
        # 7. Populate DI blocks dynamically (Rows 39 to 50)
        for idx, c in enumerate(di_countries):
            offset = idx * di_rows_per_block
            base = 39 + fdi_shift + offset
            
            new_cells[f"B{base+2}"] = c.get("country", "") # B41
            new_cells[f"C{base+2}"] = c.get("equity_percent_py", 0) # C41
            new_cells[f"D{base+2}"] = c.get("equity_percent_fy", 0) # D41
            self.sec3_pct_cells.extend([f"C{base+2}", f"D{base+2}"])

            
            new_cells[f"D{base+5}"] = c.get("equity_capital_py", 0) # D44
            new_cells[f"E{base+5}"] = c.get("equity_capital_fy", 0) # E44
            
            new_cells[f"D{base+6}"] = c.get("liabilities_py", 0) # D45
            new_cells[f"E{base+6}"] = c.get("liabilities_fy", 0) # E45
            
            new_cells[f"D{base+7}"] = 0.0 # D46 Claims
            new_cells[f"E{base+7}"] = 0.0 # E46 Claims
            
            new_cells[f"D{base+8}"] = c.get("other_capital_py", 0) # D47
            new_cells[f"E{base+8}"] = c.get("other_capital_fy", 0) # E47
            
            new_cells[f"D{base+9}"] = c.get("other_liabilities_py", 0) # D48
            new_cells[f"E{base+9}"] = c.get("other_liabilities_fy", 0) # E48
            
            new_cells[f"D{base+10}"] = c.get("other_claims_py", 0) # D49
            new_cells[f"E{base+10}"] = c.get("other_claims_fy", 0) # E49
            
            new_cells[f"D{base+11}"] = 0.0 # D50 Disinvestments
            new_cells[f"E{base+11}"] = 0.0 # E50 Disinvestments

        # 8. Shift unmerged coordinates
        sec3_unmerge = []
        for rng_str in unmerge_targets["Section III"]:
            parts = rng_str.split(":")
            match = re.match(r"^[a-zA-Z]+(\d+)", parts[0])
            row_num = int(match.group(1)) if match else 0
            
            if 39 <= row_num <= 50:
                for i in range(di_blocks_to_add + 1):
                    offset = fdi_shift + (i * di_rows_per_block)
                    new_parts = []
                    for p in parts:
                        m = re.match(r"^([a-zA-Z]+)(\d+)$", p)
                        if m: new_parts.append(f"{m.group(1)}{int(m.group(2)) + offset}")
                        else: new_parts.append(p)
                    sec3_unmerge.append(":".join(new_parts))
            elif 15 <= row_num <= 27:
                for i in range(fdi_blocks_to_add + 1):
                    offset = i * fdi_rows_per_block
                    new_parts = []
                    for p in parts:
                        m = re.match(r"^([a-zA-Z]+)(\d+)$", p)
                        if m: new_parts.append(f"{m.group(1)}{int(m.group(2)) + offset}")
                        else: new_parts.append(p)
                    sec3_unmerge.append(":".join(new_parts))
            else:
                new_parts = []
                for p in parts:
                    m = re.match(r"^([a-zA-Z]+)(\d+)$", p)
                    if m:
                        r_num = int(m.group(2))
                        if r_num > 50: new_parts.append(f"{m.group(1)}{r_num + fdi_shift + di_shift}")
                        elif r_num > 27: new_parts.append(f"{m.group(1)}{r_num + fdi_shift}")
                        else: new_parts.append(p)
                    else: new_parts.append(p)
                sec3_unmerge.append(":".join(new_parts))
        unmerge_targets["Section III"] = sec3_unmerge
        
        # 9. Shift static cells passed in
        for coord, val in cells.items():
            match = re.match(r"^([a-zA-Z]+)(\d+)$", coord)
            if match:
                col_part, row_part = match.groups()
                row_num = int(row_part)
                
                if row_num > 50:
                    new_coord = f"{col_part}{row_num + fdi_shift + di_shift}"
                elif row_num > 27:
                    new_coord = f"{col_part}{row_num + fdi_shift}"
                else:
                    new_coord = coord
                    
                # Do not let empty static defaults overwrite our dynamically populated FDI/DI values
                if new_coord in new_cells and val in ["", 0, "0", 0.0, None, "nan"]:
                    continue
                    
                new_cells[new_coord] = val
            else:
                if coord not in new_cells or val not in ["", 0, "0", 0.0, None, "nan"]:
                    new_cells[coord] = val
                
        return new_cells

