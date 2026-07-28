"""
test_dom_extraction.py — Validates the DOM pipeline by extracting key variables.

Scans all .md files in the ocr_output folder, builds a DOM for each, then
searches for each variable across ALL DOMs.  If a variable appears in
multiple files, every occurrence is reported (with the source file name).

Exports results to an Excel file with columns:
  Source File, Variable, Status, Current Year Value, Prior Year Value, …

Usage:
    cd automation_engine/modules/idp_studio
    python -m dom_learner.test_dom_extraction
"""

from __future__ import annotations

import csv
import json
import logging
from collections import defaultdict
import sys
import time
from dataclasses import dataclass
from pathlib import Path

_THIS_DIR = Path(__file__).resolve().parent
_IDP_STUDIO_DIR = _THIS_DIR.parent
if str(_IDP_STUDIO_DIR) not in sys.path:
    sys.path.insert(0, str(_IDP_STUDIO_DIR))

from dom_learner.engine.dom_builder import DOMBuilder
from dom_learner.engine.dom_query import DOMQuery
from dom_learner.models import DOMNode, NodeType

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

_OCR_OUTPUT_DIR = _THIS_DIR / "Docs" / "ocr_output"
_DEFAULT_OUTPUT = _THIS_DIR / "output" / "dom_extraction_test.xlsx"

# Variables to extract — single flat list, searched across ALL DOMs.
# These can come from any .md file; we don't pre-map them to files.
VARIABLES = [
    # Balance sheet items
    "Partners' contribution",
    "Partners' current account",
    "Trade payables",
    "Other current liabilities",
    "Short-term provisions",
    "Trade receivables",
    "Cash and bank balances",
    "Revenue from operations",
    "Employee benefits expense",
    "Professional & Consultancy Charges",
    # Additional items that may appear in other documents
    "Reserves and surplus",
    "Profit before tax",
    "Long-term provisions",
    "Other long-term liabilities",
    "Tangible assets",
]


# ---------------------------------------------------------------------------
# Extraction result
# ---------------------------------------------------------------------------

@dataclass
class ExtractionResult:
    variable: str
    source_file: str  # which .md file this was found in
    dom_path: str
    depth: int
    node_id: str
    section: str
    table_id: str
    current_year_value: str
    prior_year_value: str
    status: str  # FOUND / PARTIAL / NOT_FOUND


# ---------------------------------------------------------------------------
# Smart variable extractor
# ---------------------------------------------------------------------------

def _looks_like_number(val: str) -> bool:
    """Return True if val looks like a financial number."""
    cleaned = val.replace(",", "").replace(".", "").replace("-", "").strip()
    return bool(cleaned and cleaned.isdigit())


def _find_row_by_label(q: DOMQuery, label: str) -> DOMNode | None:
    """Find a row whose cells contain *label* (case-insensitive substring).

    The standard `q.find_row()` uses row_label metadata which is the first
    cell — but in this document the first cell is often just "(a)", "(b)", etc.
    So we also search all cells within rows for the label text.
    
    If multiple rows match, we use a scoring heuristic to pick the best one.
    This helps bypass OCR artifacts in the main tables by preferring the 
    cleaner detailed note tables.
    """
    label_lower = label.strip().lower()

    # Get all candidate rows
    matches = q.find_row(label)
    if not matches:
        cell_matches = q.find_by_text(label, node_type=NodeType.CELL)
        matches = [c.parent for c in cell_matches if c.parent and c.parent.node_type == NodeType.ROW]

    if not matches:
        return None

    best_row = None
    best_score = -999

    for row in matches:
        fy25, fy24 = _get_value_cells(row)
        score = 0

        # Prefer rows that successfully extracted two values
        if fy25 and fy24:
            score += 10
            
            # Penalize non-numeric values heavily
            if not _looks_like_number(fy25) or not _looks_like_number(fy24):
                score -= 100
            
            # Penalize values that are actually date headers from note tables
            if "march" in fy25.lower() or "march" in fy24.lower():
                score -= 50
                
            # Penalize OCR artifact values (e.g. dots instead of commas)
            if "." in fy25 or "." in fy24:
                score -= 5

        # Prefer exact matches or aggregated totals
        row_text = row.text.lower()
        if label_lower == row_text or f"| {label_lower}" in row_text:
            score += 5
        elif f"total {label_lower}" in row_text:
            score += 8
        elif f"{label_lower} (net)" in row_text:
            score += 8

        if score > best_score:
            best_score = score
            best_row = row

    return best_row


def _get_value_cells(row: DOMNode) -> tuple[str, str]:
    """Extract current-year and prior-year values from a row.

    Strategy: the last two cells with numeric-looking content are typically
    the current year and previous year values.  We also look at column
    headers containing year indicators.
    """
    cells = [c for c in row.children if c.node_type == NodeType.CELL]

    current_year = ""
    prior_year = ""

    # Strategy 1: look for cells whose column header contains year keywords
    # Try multiple year patterns to handle different documents
    year_pairs = [
        ("2025", "2024"),
        ("2022", "2021"),
        ("2023", "2022"),
        ("2024", "2023"),
    ]

    for curr_yr, prev_yr in year_pairs:
        for cell in cells:
            col_header = cell.metadata.get("column_header", "").lower()
            text = cell.text.strip()
            if not text or text in ("", "(empty)"):
                continue

            if curr_yr in col_header:
                current_year = text
            elif prev_yr in col_header:
                prior_year = text

        if current_year or prior_year:
            break

    # Strategy 2: if headers don't have year, use positional (last two numeric cells)
    if not current_year and not prior_year:
        numeric_cells = []
        for cell in cells:
            text = cell.text.strip()
            # check if text looks numeric (contains digits and commas)
            cleaned = text.replace(",", "").replace(".", "").replace("-", "").replace(" ", "")
            if cleaned and any(c.isdigit() for c in cleaned):
                numeric_cells.append(cell)

        if len(numeric_cells) >= 2:
            current_year = numeric_cells[-2].text.strip()
            prior_year = numeric_cells[-1].text.strip()
        elif len(numeric_cells) == 1:
            current_year = numeric_cells[0].text.strip()

    return current_year, prior_year


def _format_dom_path(q: DOMQuery, node: DOMNode) -> str:
    """Format the structural path for display."""
    path = q.get_structural_path(node)
    parts = []
    for p in path:
        label = p.get("label") or p.get("text") or p["type"]
        # truncate long labels
        if len(label) > 40:
            label = label[:37] + "..."
        parts.append(f"{p['type']}({label})")
    return " → ".join(parts)


def extract_variables(
    dom_queries: dict[str, DOMQuery],
) -> list[ExtractionResult]:
    """Search for all variables across all DOMs.

    Parameters
    ----------
    dom_queries : dict[str, DOMQuery]
        Mapping of ``filename → DOMQuery`` for every .md file.

    Returns
    -------
    list[ExtractionResult]
        One result per (variable, source_file) match.  If a variable is
        found in zero files a single NOT_FOUND row is emitted.
    """
    results: list[ExtractionResult] = []

    for var_name in VARIABLES:
        found_in_any = False

        for filename, q in dom_queries.items():
            row = _find_row_by_label(q, var_name)

            if row is None:
                continue

            found_in_any = True

            # Get values
            curr_val, prev_val = _get_value_cells(row)

            # Get structural context
            dom_path = _format_dom_path(q, row)
            section_node = row.section
            table_node = row.table

            status = "FOUND" if curr_val else "PARTIAL"

            results.append(ExtractionResult(
                variable=var_name,
                source_file=filename,
                dom_path=dom_path,
                depth=row.depth,
                node_id=row.id,
                section=section_node.text[:60] if section_node else "—",
                table_id=table_node.id if table_node else "—",
                current_year_value=curr_val,
                prior_year_value=prev_val,
                status=status,
            ))

        if not found_in_any:
            results.append(ExtractionResult(
                variable=var_name,
                source_file="—",
                dom_path="—",
                depth=0,
                node_id="—",
                section="—",
                table_id="—",
                current_year_value="—",
                prior_year_value="—",
                status="NOT_FOUND",
            ))

    return results


# ---------------------------------------------------------------------------
# Duplicate detection
# ---------------------------------------------------------------------------

def _report_duplicates(
    results: list[ExtractionResult],
    log: logging.Logger,
) -> dict[str, list[str]]:
    """Detect variables found in multiple files and log them.

    Returns a dict mapping variable name → list of source files.
    """
    var_files: dict[str, list[str]] = defaultdict(list)
    for r in results:
        if r.status != "NOT_FOUND":
            var_files[r.variable].append(r.source_file)

    duplicates = {v: files for v, files in var_files.items() if len(files) > 1}

    if duplicates:
        log.info("")
        log.info("⚠️  Duplicate variables found across files:")
        for var, files in duplicates.items():
            file_list = ", ".join(files)
            log.info("    %-40s → %s", var, file_list)
    else:
        log.info("")
        log.info("✅  No duplicate variables across files.")

    return duplicates


# ---------------------------------------------------------------------------
# Excel export (using openpyxl if available, falls back to CSV)
# ---------------------------------------------------------------------------

def _export_excel(results: list[ExtractionResult], output_path: Path) -> Path:
    """Export results to Excel. Falls back to CSV if openpyxl is not installed."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        return _export_xlsx(results, output_path)
    except ImportError:
        # fallback to CSV
        csv_path = output_path.with_suffix(".csv")
        return _export_csv(results, csv_path)


def _export_xlsx(results: list[ExtractionResult], output_path: Path) -> Path:
    """Export to a styled Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DOM Extraction Test"

    # --- Styles ---
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    found_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    not_found_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    dup_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # --- Title row ---
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "DOM Learning Engine — Multi-File Extraction Test Results"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center")

    # --- Blank row ---
    ws.append([])

    # --- Headers ---
    headers = [
        "Source File",
        "Variable",
        "Status",
        "Current Year Value",
        "Prior Year Value",
        "DOM Depth",
        "Node ID",
        "Section",
        "Table ID",
        "DOM Path / Location",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # --- Detect duplicates for highlighting ---
    dup_vars: set[str] = set()
    var_counts: dict[str, int] = defaultdict(int)
    for r in results:
        if r.status != "NOT_FOUND":
            var_counts[r.variable] += 1
    dup_vars = {v for v, c in var_counts.items() if c > 1}

    # --- Data rows ---
    for i, r in enumerate(results, start=4):
        row_data = [
            r.source_file,
            r.variable,
            r.status,
            r.current_year_value,
            r.prior_year_value,
            r.depth,
            r.node_id,
            r.section,
            r.table_id,
            r.dom_path,
        ]
        ws.append(row_data)

        # status coloring
        status_cell = ws.cell(row=i, column=3)
        if r.status == "FOUND":
            status_cell.fill = found_fill
        elif r.status == "PARTIAL":
            status_cell.fill = partial_fill
        else:
            status_cell.fill = not_found_fill

        # highlight duplicates
        if r.variable in dup_vars and r.status != "NOT_FOUND":
            ws.cell(row=i, column=2).fill = dup_fill

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Summary row ---
    ws.append([])
    summary_row = len(results) + 5
    found = sum(1 for r in results if r.status == "FOUND")
    partial = sum(1 for r in results if r.status == "PARTIAL")
    not_found = sum(1 for r in results if r.status == "NOT_FOUND")

    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value=f"Found: {found}/{len(results)}")
    ws.cell(row=summary_row + 1, column=1).fill = found_fill
    ws.cell(row=summary_row + 2, column=1, value=f"Partial: {partial}/{len(results)}")
    ws.cell(row=summary_row + 2, column=1).fill = partial_fill
    ws.cell(row=summary_row + 3, column=1, value=f"Not Found: {not_found}/{len(results)}")
    ws.cell(row=summary_row + 3, column=1).fill = not_found_fill
    ws.cell(row=summary_row + 4, column=1, value=f"Duplicates: {len(dup_vars)} variable(s)")
    ws.cell(row=summary_row + 4, column=1).fill = dup_fill

    # --- Per-file breakdown ---
    file_stats_row = summary_row + 6
    ws.cell(row=file_stats_row, column=1, value="Per-File Breakdown").font = Font(bold=True)
    source_files = sorted({r.source_file for r in results if r.source_file != "—"})
    for idx, sf in enumerate(source_files):
        sf_found = sum(1 for r in results if r.source_file == sf and r.status == "FOUND")
        sf_partial = sum(1 for r in results if r.source_file == sf and r.status == "PARTIAL")
        sf_total = sum(1 for r in results if r.source_file == sf)
        ws.cell(
            row=file_stats_row + 1 + idx, column=1,
            value=f"{sf}: {sf_found} found, {sf_partial} partial (of {sf_total} matches)",
        )

    # --- Column widths ---
    col_widths = [35, 35, 12, 18, 18, 10, 12, 40, 12, 80]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # --- Save ---
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def _export_csv(results: list[ExtractionResult], output_path: Path) -> Path:
    """Fallback CSV export."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Source File", "Variable", "Status",
            "Current Year Value", "Prior Year Value",
            "DOM Depth", "Node ID", "Section", "Table ID", "DOM Path",
        ])
        for r in results:
            writer.writerow([
                r.source_file, r.variable, r.status,
                r.current_year_value, r.prior_year_value,
                r.depth, r.node_id, r.section, r.table_id, r.dom_path,
            ])
    return output_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%H:%M:%S",
    )
    log = logging.getLogger("dom_learner.test")

    log.info("=" * 70)
    log.info("   DOM Extraction Test — Multi-File (%d Variables)", len(VARIABLES))
    log.info("=" * 70)

    # --- Discover all .md files ---
    md_files = sorted(_OCR_OUTPUT_DIR.glob("*.md"))
    if not md_files:
        log.error("No .md files found in %s", _OCR_OUTPUT_DIR)
        sys.exit(1)

    log.info("Found %d .md file(s) in %s:", len(md_files), _OCR_OUTPUT_DIR.name)
    for f in md_files:
        log.info("    • %s", f.name)

    # --- Build DOM for each file ---
    dom_queries: dict[str, DOMQuery] = {}
    builder = DOMBuilder()

    for md_file in md_files:
        log.info("")
        log.info("Building DOM from %s...", md_file.name)
        t0 = time.perf_counter()
        document = builder.build(md_file)
        q = DOMQuery(document)
        dom_queries[md_file.name] = q
        log.info(
            "  DOM built in %.3fs (%d nodes)",
            time.perf_counter() - t0,
            sum(q.summary().values()),
        )

        # --- Save DOM tree to JSON ---
        _DEFAULT_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
        dom_json_path = _DEFAULT_OUTPUT.parent / f"{md_file.stem}.dom.json"
        with dom_json_path.open("w", encoding="utf-8") as f:
            json.dump(document.to_dict(), f, indent=2)
        log.info("  Saved DOM tree to %s", dom_json_path.name)

    # --- Extract variables across all DOMs ---
    log.info("")
    log.info("Extracting %d variables across %d DOM(s)...", len(VARIABLES), len(dom_queries))
    log.info("-" * 70)
    results = extract_variables(dom_queries)

    # --- Print results to console ---
    for r in results:
        icon = "✅" if r.status == "FOUND" else ("⚠️" if r.status == "PARTIAL" else "❌")
        src = r.source_file if r.source_file != "—" else "(none)"
        log.info(
            "  %s %-40s CY=%-15s PY=%-15s [%s] ← %s",
            icon, r.variable,
            r.current_year_value or "—",
            r.prior_year_value or "—",
            r.node_id, src,
        )

    # --- Duplicate report ---
    _report_duplicates(results, log)

    # --- Summary ---
    found = sum(1 for r in results if r.status == "FOUND")
    partial = sum(1 for r in results if r.status == "PARTIAL")
    not_found = sum(1 for r in results if r.status == "NOT_FOUND")

    log.info("-" * 70)
    log.info(
        "  ✅ Found: %d/%d   ⚠️ Partial: %d/%d   ❌ Not Found: %d/%d",
        found, len(results), partial, len(results), not_found, len(results),
    )

    # --- Export to Excel ---
    log.info("")
    output_path = _export_excel(results, _DEFAULT_OUTPUT)
    log.info("Exported results → %s", output_path)
    log.info("=" * 70)


if __name__ == "__main__":
    main()
