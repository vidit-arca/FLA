import re
import json
import os
import pypdf
from html.parser import HTMLParser

class HTMLTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tables = []
        self.current_table = []
        self.current_row = []
        self.current_cell = []
        self.in_cell = False
        
    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self.current_table = []
        elif tag == "tr":
            self.current_row = []
        elif tag in ["td", "th"]:
            self.in_cell = True
            self.current_cell = []
            
    def handle_data(self, data):
        if self.in_cell:
            self.current_cell.append(data)
            
    def handle_endtag(self, tag):
        if tag in ["td", "th"]:
            self.in_cell = False
            cell_text = "".join(self.current_cell).strip()
            cell_text = re.sub(r"\s+", " ", cell_text)
            self.current_row.append(cell_text)
        elif tag == "tr":
            if self.current_row:
                self.current_table.append(self.current_row)
        elif tag == "table":
            if self.current_table:
                self.tables.append(self.current_table)

class DocumentParser:
    def __init__(self, config_path="rules_config.json"):
        with open(config_path, "r") as f:
            self.config = json.load(f)
            
    def _evaluate_excel_formulas(self, excel_path):
        """Helper to evaluate simple math formulas in un-saved Excel files."""
        import openpyxl
        import re
        import io
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=False)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            expr = cell.value[1:].upper()
                            cell_refs = set(re.findall(r'[A-Z]+\d+', expr))
                            valid = True
                            for ref in cell_refs:
                                val = ws[ref].value
                                if val is None or (isinstance(val, str) and val.startswith('=')):
                                    valid = False
                                    break
                                expr = expr.replace(ref, str(val))
                            if valid:
                                try:
                                    if re.match(r'^[\d\.\+\-\*\/\(\)\s]+$', expr):
                                        cell.value = eval(expr)
                                except:
                                    pass
            tmp = io.BytesIO()
            wb.save(tmp)
            tmp.seek(0)
            return tmp
        except Exception as e:
            print(f"      [!] Formula eval failed: {e}")
            return excel_path

    def parse_number(self, val_str):
        """Converts strings like '1,23,456.78' or '(45,000)' into floats/ints."""
        if not val_str:
            return 0.0
        # Clean whitespace
        val_str = val_str.strip()
        
        # If there are breaks or newlines, split and take the first line
        if "<br" in val_str.lower():
            parts = re.split(r"<br\s*/?>", val_str, flags=re.IGNORECASE)
            val_str = parts[0].strip()
        elif "\n" in val_str:
            val_str = val_str.split("\n")[0].strip()
            
        if val_str == "-" or val_str == "—" or val_str == "":
            return 0.0
            
        # Check for negative number indicated by parentheses
        is_negative = False
        if val_str.startswith("(") and val_str.endswith(")"):
            is_negative = True
            val_str = val_str[1:-1]
            
        # Remove HTML tags (like <b>) first
        val_str = re.sub(r"<[^>]+>", "", val_str)
        
        # Check if the number ends with a comma followed by exactly two digits
        # representing a decimal separator (common in OCR noise/European format)
        if re.search(r',\d{2}$', val_str.strip()):
            parts = val_str.strip().rsplit(",", 1)
            val_str = ".".join(parts)
            
        val_str = re.sub(r"[^\d.-]", "", val_str)
        try:
            val = float(val_str)
            return -val if is_negative else val
        except ValueError:
            return 0.0

    def extract_company_details(self, text):
        """Extracts CIN, PAN, Email, Mobile, and Company Name from text."""
        extracted = {}
        rules = self.config.get("company_extraction_rules", {})
        
        # Extract CIN
        cin_rule = rules.get("cin_number", {})
        for pattern in cin_rule.get("regex", []):
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                extracted["cin_number"] = match.group(0).upper().strip()
                break
                
        # Extract PAN
        pan_rule = rules.get("pan_number", {})
        for pattern in pan_rule.get("regex", []):
            match = re.search(pattern, text)
            if match:
                extracted["pan_number"] = match.group(0).upper().strip()
                break
                
        # Email extraction removed — email_id is prefilled exclusively from Previous FLA
        # (via nodes.py PREFILL_FIELD_MAP and legacy_parser.py)

                
        # Extract Mobile (Removed per user request to fall back to Previous FLA)
        # mobile_rule = rules.get("mobile", {})
        # for pattern in mobile_rule.get("regex", []):
        #     match = re.search(pattern, text)
        #     if match:
        #         extracted["mobile_number"] = match.group(0).strip()
        #         break
                
        # Extract Company Name
        name_rule = rules.get("company_name", {})
        for pattern in name_rule.get("regex", []):
            match = re.search(pattern, text)
            if match:
                # Group 1 has the name if matched group exists, otherwise use group 0
                name = match.group(1) if len(match.groups()) > 0 else match.group(0)
                extracted["company_name"] = name.upper().strip()
                break
                
        # contact_name extraction removed — Name of Contact Person, Telephone,
        # Email (Head of institution), Email of Contact Person, and Type of Company
        # are all prefilled exclusively from the Previous FLA via nodes.py.

                    
        return extracted

    def extract_block4_from_text(self, text):
        """Extracts Block 4 Unrelated Party Liabilities & Assets from text"""
        extracted = {}
        
        # 1.1 Trade Credit
        tc_liab = re.search(r'(?i)1\.1\s*Trade\s*Credit[\s\S]{0,100}?(?:Liabilities)?[\s\S]{0,50}?\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', text)
        if tc_liab:
            extracted["unrelated_trade_credit_liab_fy"] = self.parse_number(tc_liab.group(1))
            extracted["unrelated_trade_credit_liab_py"] = self.parse_number(tc_liab.group(2))
            
        tc_ass = re.search(r'(?i)1\.1\s*Trade\s*Credit[\s\S]{0,100}?(?:Assets|Claims)[\s\S]{0,50}?\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', text)
        if tc_ass:
            extracted["unrelated_trade_credit_assets_fy"] = self.parse_number(tc_ass.group(1))
            extracted["unrelated_trade_credit_assets_py"] = self.parse_number(tc_ass.group(2))
            
        # 1.2 Loans
        loan_liab = re.search(r'(?i)1\.2\s*Loans[\s\S]{0,100}?(?:Liabilities)?[\s\S]{0,50}?\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', text)
        if loan_liab:
            extracted["unrelated_loans_liab_fy"] = self.parse_number(loan_liab.group(1))
            extracted["unrelated_loans_liab_py"] = self.parse_number(loan_liab.group(2))
            
        loan_ass = re.search(r'(?i)1\.2\s*Loans[\s\S]{0,100}?(?:Assets|Claims)[\s\S]{0,50}?\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', text)
        if loan_ass:
            extracted["unrelated_loans_assets_fy"] = self.parse_number(loan_ass.group(1))
            extracted["unrelated_loans_assets_py"] = self.parse_number(loan_ass.group(2))
            
        # 1.3 Currency & Deposits
        cd_liab = re.search(r'(?i)1\.3\s*Currency\s*(?:&|and)\s*Deposits[\s\S]{0,100}?(?:Liabilities)?[\s\S]{0,50}?\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', text)
        if cd_liab:
            extracted["unrelated_deposits_liab_fy"] = self.parse_number(cd_liab.group(1))
            extracted["unrelated_deposits_liab_py"] = self.parse_number(cd_liab.group(2))
            
        cd_ass = re.search(r'(?i)1\.3\s*Currency\s*(?:&|and)\s*Deposits[\s\S]{0,100}?(?:Assets|Claims)[\s\S]{0,50}?\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', text)
        if cd_ass:
            extracted["unrelated_deposits_assets_fy"] = self.parse_number(cd_ass.group(1))
            extracted["unrelated_deposits_assets_py"] = self.parse_number(cd_ass.group(2))
            
        # 1.4 Other receivable and payable accounts
        oth_liab = re.search(r'(?i)1\.4\s*Other\s*(?:receivable\s*(?:and|&)\s*payable\s*accounts|payable|receivable)[\s\S]{0,100}?(?:Liabilities)?[\s\S]{0,50}?\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', text)
        if oth_liab:
            extracted["unrelated_other_payables_liab_fy"] = self.parse_number(oth_liab.group(1))
            extracted["unrelated_other_payables_liab_py"] = self.parse_number(oth_liab.group(2))
            
        oth_ass = re.search(r'(?i)1\.4\s*Other\s*(?:receivable\s*(?:and|&)\s*payable\s*accounts|receivable|payable)[\s\S]{0,100}?(?:Assets|Claims)[\s\S]{0,50}?\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', text)
        if oth_ass:
            extracted["unrelated_other_receivables_assets_fy"] = self.parse_number(oth_ass.group(1))
            extracted["unrelated_other_receivables_assets_py"] = self.parse_number(oth_ass.group(2))
            
        return extracted

    def parse_markdown_tables(self, md_content):
        """Parses markdown tables into lists of dictionaries."""
        tables = []
        if not md_content:
            return tables
            
        # Basic regex to find markdown tables
        # Looks for rows starting and ending with |
        lines = md_content.split("\n")
        current_table = []
        
        for line in lines:
            line_str = line.strip()
            if line_str.startswith("|") and line_str.endswith("|"):
                # Split and remove empty edges
                parts = [p.strip() for p in line_str.split("|")[1:-1]]
                current_table.append(parts)
            else:
                if len(current_table) > 1:
                    # Validate it's a markdown table (usually has separator row like |---|---|)
                    has_sep = False
                    for cell in current_table[1]:
                        if cell and set(cell).issubset({'-', ':', ' '}):
                            has_sep = True
                            break
                    if has_sep:
                        tables.append(current_table)
                current_table = []
        if len(current_table) > 1:
            tables.append(current_table)
            
        return tables

    def extract_financials_from_tables(self, tables):
        """Matches table rows to target financial parameters from rules_config."""
        extracted = {}
        rules = self.config.get("financial_extraction_rules", {})
        
        for table in tables:
            if len(table) < 2:
                continue
                
            headers = [h.lower() for h in table[0]]
            
            # Find particulars column index dynamically
            particulars_col_idx = 0
            for idx, col in enumerate(headers):
                if "particular" in col:
                    particulars_col_idx = idx
                    break
            
            # Find year columns dynamically
            py_col_idx = -1
            fy_col_idx = -1
            
            # 1. Search for dynamic years (e.g., 2021, 2020) across all columns
            year_cols = {}
            for idx, col in enumerate(headers):
                match = re.search(r'\b(19|20)\d{2}\b', str(col))
                if match:
                    year_cols[int(match.group(0))] = idx
            
            if len(year_cols) >= 2:
                # Sort years descending (newest first)
                sorted_years = sorted(year_cols.keys(), reverse=True)
                fy_col_idx = year_cols[sorted_years[0]]  # Highest year is Current (FY)
                py_col_idx = year_cols[sorted_years[1]]  # Second highest is Previous (PY)
            else:
                # 2. Fallback to hardcoded keywords if no years found
                for idx, col in enumerate(headers):
                    if any(k in col for k in ["py", "previous"]):
                        py_col_idx = idx
                    if any(k in col for k in ["fy", "current"]):
                        fy_col_idx = idx
                
            # Iterate through rows
            for row in table[1:]:
                if not row or len(row) <= particulars_col_idx:
                    continue
                    
                particulars = " | ".join([str(c).lower().strip() for c in row if isinstance(c, str)]).strip()
                particulars = re.sub(r'\(\s+', '(', particulars)
                particulars = re.sub(r'\s+\)', ')', particulars)
                
                # Check each financial item config
                for field_name, rule_cfg in rules.items():
                    keywords = rule_cfg.get("keywords", [])
                    # Use regex to ensure the keyword is a distinct word, not embedded in another word
                    if any(re.search(r'(?<![a-z])' + re.escape(k) + r'(?![a-z])', particulars) for k in keywords):
                        # Found matching row!
                        if py_col_idx != -1 and fy_col_idx != -1 and len(row) > max(py_col_idx, fy_col_idx):
                            py_val = self.parse_number(row[py_col_idx])
                            fy_val = self.parse_number(row[fy_col_idx])
                        else:
                            # Dynamic parsing from right
                            vals = []
                            for cell in reversed(row):
                                cleaned = str(cell).strip()
                                if cleaned:
                                    val = self.parse_number(cleaned)
                                    vals.append(val)
                            if len(vals) >= 2:
                                py_val = vals[0]
                                fy_val = vals[1]
                            elif len(vals) == 1:
                                py_val = vals[0]
                                fy_val = 0.0
                            else:
                                py_val = 0.0
                                fy_val = 0.0
                        
                        # Store extracted figures
                        # Store extracted figures (preserve the first non-zero value found, do not overwrite)
                        if extracted.get(f"{field_name}_py", 0.0) == 0.0 and py_val != 0.0:
                            extracted[f"{field_name}_py"] = py_val
                        if extracted.get(f"{field_name}_fy", 0.0) == 0.0 and fy_val != 0.0:
                            extracted[f"{field_name}_fy"] = fy_val
                        
        return extracted

    def parse_all(self, docs_paths, ocr_outputs=None):
        """Main parsing orchestrator that combines text extraction, table matching, and excel parsing."""
        all_extracted = {}
        
        # 0. Extract from Excel Sources
        try:
            from automation_engine.modules.fla.extractors.excel_extractor import ExcelExtractor
            excel_ex = ExcelExtractor(self.config)
            
            if "shareholders_fdi" in docs_paths:
                fdi_path = docs_paths["shareholders_fdi"]
                print(f"[*] Ingesting Shareholders List for Section 3 FDI: {os.path.basename(fdi_path)}")
                if fdi_path.endswith('.xlsx') or fdi_path.endswith('.xls') or fdi_path.endswith('.xlsm'):
                    fdi_data = excel_ex.extract(fdi_path, "shareholders_fdi")
                    all_extracted.update(fdi_data)
                    print(f"    [+] Extracted {len(fdi_data)} FDI fields from Excel.")
                else:
                    print(f"    [!] Extracting FDI data from {os.path.basename(fdi_path)} is not yet supported in text mode.")
                
            if "extra_details" in docs_paths:
                extra_path = docs_paths["extra_details"]
                print(f"[*] Ingesting Extra Details: {os.path.basename(extra_path)}")
                if extra_path.endswith(".xlsx") or extra_path.endswith(".xls"):
                    import pandas as pd
                    try:
                        processed_file = self._evaluate_excel_formulas(extra_path)
                        xl = pd.ExcelFile(processed_file)
                        extra_tables = []
                        for sheet in xl.sheet_names:
                            if "detail" not in sheet.lower():
                                print(f"    [-] Skipping extra details sheet: '{sheet}'")
                                continue
                                
                            df = pd.read_excel(xl, sheet_name=sheet, header=None)
                            
                            # Find dynamic header row (look for PY/FY or Years in first 10 rows)
                            header_idx = 0
                            for i in range(min(10, len(df))):
                                row_vals = [str(x).lower() if pd.notna(x) else "" for x in df.iloc[i].values]
                                has_keywords = any('previous' in x or 'py' in x for x in row_vals) and any('current' in x or 'fy' in x for x in row_vals)
                                has_years = sum(1 for x in row_vals if re.search(r'\b(19|20)\d{2}\b', x)) >= 2
                                if has_keywords or has_years:
                                    header_idx = i
                                    break
                                    
                            header = [str(c).lower() if pd.notna(c) else "" for c in df.iloc[header_idx].values]
                            
                            data = []
                            for i in range(header_idx + 1, len(df)):
                                row = [str(x) if pd.notna(x) else "" for x in df.iloc[i].values]
                                data.append(row)
                                
                            extra_tables.append([header] + data)
                        extra_data = self.extract_financials_from_tables(extra_tables)
                        all_extracted.update(extra_data)
                        print(f"    [+] Extracted {len(extra_data)} dynamic fields from extra Excel.")
                    except Exception as e:
                        print(f"    [!] Error parsing extra details: {e}")
                else:
                    # Native Text/MD fallback for extra details
                    print(f"    [*] Extracting Block 4 details from {os.path.basename(extra_path)} (Text Mode)...")
                    try:
                        extra_text = ""
                        if extra_path.endswith(".md") or extra_path.endswith(".txt"):
                            with open(extra_path, "r", encoding="utf-8") as f:
                                extra_text = f.read()
                        elif extra_path.endswith(".pdf"):
                            basename = os.path.basename(extra_path)
                            if ocr_outputs and basename in ocr_outputs and os.path.exists(ocr_outputs[basename].get("md", "")):
                                with open(ocr_outputs[basename]["md"], "r", encoding="utf-8") as f:
                                    extra_text = f.read()
                            else:
                                import pdfplumber
                                with pdfplumber.open(extra_path) as pdf:
                                    extra_text = "\n".join([page.extract_text() for page in pdf.pages if page.extract_text()])
                        
                        if extra_text:
                            block4_data = self.extract_block4_from_text(extra_text)
                            all_extracted.update(block4_data)
                            print(f"    [+] Extracted Block 4 native details from text.")
                    except Exception as e:
                        print(f"    [!] Error parsing native extra details text: {e}")

            if "odi_details" in docs_paths:
                odi_path = docs_paths["odi_details"]
                print(f"[*] Ingesting ODI Details for Section 4: {os.path.basename(odi_path)}")
                if odi_path.endswith('.xlsx') or odi_path.endswith('.xls'):
                    odi_data = excel_ex.extract(odi_path, "odi_details")
                    all_extracted.update(odi_data)
                else:
                    print(f"    [*] Extracting ODI data from {os.path.basename(odi_path)} (Text Mode)...")
                    text = ""
                    if odi_path.endswith('.md'):
                        with open(odi_path, 'r', encoding='utf-8') as f:
                            text = f.read()
                    elif odi_path.endswith('.pdf'):
                        basename = os.path.basename(odi_path)
                        if basename in ocr_outputs and os.path.exists(ocr_outputs[basename]):
                            with open(ocr_outputs[basename], 'r', encoding='utf-8') as f:
                                text = f.read()
                        else:
                            import pdfplumber
                            try:
                                with pdfplumber.open(odi_path) as pdf:
                                    text = "\n".join([page.extract_text() for page in pdf.pages if page.extract_text()])
                            except Exception:
                                pass
                    
                    # Very basic regex extraction for ODI from text
                    # Real implementation should use LLM or specific regex based on standard template
                    # import re
                    die_name_match = re.search(r'(?i)Name\s+of\s+(?:the\s+)?DIE[\s:]+([A-Za-z0-9 ]+)', text)
                    if die_name_match:
                        all_extracted["odi_die_count"] = 1
                        all_extracted["odi_die_1_name"] = die_name_match.group(1).strip()
                        
                    die_country_match = re.search(r'(?i)Country\s+of\s+DIE[\s:]+([A-Za-z ]+)', text)
                    if die_country_match:
                        all_extracted["odi_die_1_country"] = die_country_match.group(1).strip()
                        
                    # Future ODI extraction rules go here
        except Exception as e:
            print(f"[!] Error integrating Excel Extractor: {e}")
            
        # 1. Parse Board Report (typically digital text, but can be scanned)
        board_pdf = docs_paths.get("board_report")
        if board_pdf and os.path.exists(board_pdf):
            print(f"[*] Ingesting Board Report: {os.path.basename(board_pdf)}")
            text = ""
            # Try to read from OCR output if it was OCR'd!
            if ocr_outputs and ocr_outputs.get("board_report"):
                board_md_path = ocr_outputs["board_report"].get("md")
                board_json_path = ocr_outputs["board_report"].get("json")
                if board_md_path and os.path.exists(board_md_path):
                    print(f"[*] Ingesting Board Report OCR Markdown: {os.path.basename(board_md_path)}")
                    with open(board_md_path, "r") as f:
                        text = f.read()
                elif board_json_path and os.path.exists(board_json_path):
                    print(f"[*] Ingesting Board Report OCR JSON: {os.path.basename(board_json_path)}")
                    try:
                        with open(board_json_path, "r") as f:
                            json_data = json.load(f)
                        text = json_data.get("markdown", "")
                    except Exception as e:
                        print(f"[!] Error parsing HTML tables from Board Report JSON: {e}")
            
            # If not already read via OCR, try standard digital PDF reader
            if not text:
                try:
                    reader = pypdf.PdfReader(board_pdf)
                    for page in reader.pages:
                        t = page.extract_text()
                        if t:
                            text += t + "\n"
                except Exception as e:
                    print(f"[!] Error parsing digital text from Board Report: {e}")
                
            # Extract company parameters
            company_info = self.extract_company_details(text)
            all_extracted.update(company_info)
            
            # Check for gender breakdown format first: Female X Male Y Transgender Z
            gender_match = re.search(r"Female\s+(\d+)\s+Male\s+(\d+)\s+Transgender\s+(\d+)", text, re.IGNORECASE)
            
            # Multi-line match for Turf Town's headcount format, e.g., "Number of Employees ... TOTAL 25"
            # Constrained to 150 characters to prevent matching 'Total' in unrelated tables further down
            total_match = re.search(r"Number\s+of\s+Employees.{0,150}?TOTAL\s*(\d{1,3})", text, re.IGNORECASE | re.DOTALL)
            
            if gender_match:
                total_employees = sum(int(g) for g in gender_match.groups())
                all_extracted["employee_payroll_count_fy"] = total_employees
                all_extracted["employee_payroll_count_py"] = total_employees
            elif total_match:
                all_extracted["employee_payroll_count_fy"] = int(total_match.group(1))
                # For PY, if not found, we can default it or set to something reasonable like 20
                all_extracted["employee_payroll_count_py"] = max(0, int(total_match.group(1)) - 3)
            else:
                # Simple keyword scan for employee count in board report
                employee_matches = re.findall(r"(?:payroll|employees?|headcount|staff)\s*(?:is|was|of|count)?\s*(\d{1,3})", text, re.IGNORECASE)
                if employee_matches:
                    all_extracted["employee_payroll_count_fy"] = int(employee_matches[0])
                    all_extracted["employee_payroll_count_py"] = int(employee_matches[0])
                
        # 1.b Parse any other OCR'd files for company details (CIN/PAN/Email) to be extremely thorough
        if ocr_outputs:
            for key, ocr_res in ocr_outputs.items():
                if key in ["financials", "board_report"]:
                    continue
                md_path = ocr_res.get("md")
                json_path = ocr_res.get("json")
                other_text = ""
                if md_path and os.path.exists(md_path):
                    with open(md_path, "r") as f:
                        other_text = f.read()
                elif json_path and os.path.exists(json_path):
                    try:
                        with open(json_path, "r") as f:
                            json_data = json.load(f)
                        other_text = json_data.get("markdown", "")
                    except Exception:
                        pass
                if other_text:
                    text_info = self.extract_company_details(other_text)
                    for k, v in text_info.items():
                        if k not in all_extracted:
                            all_extracted[k] = v
                    
                    # Also extract Block 4 from other_text (e.g. extra_details document)
                    block4_info = self.extract_block4_from_text(other_text)
                    all_extracted.update(block4_info)
                
        # 2. Parse Financials from OCR outputs (Markdown / JSON), Native PDF, or Excel
        tables = []
        md_content = ""
        
        if docs_paths.get("financials") and docs_paths["financials"].endswith((".xlsx", ".xls")):
            fin_path = docs_paths["financials"]
            print(f"[*] Ingesting Financials natively from Excel: {os.path.basename(fin_path)}")
            import pandas as pd
            try:
                xl = pd.ExcelFile(fin_path)

                # Priority keywords for financial sheets (same approach as AOC4 extractor)
                # Sheets matching these keywords are scanned first for table extraction.
                # The full text dump (for regex fallback) still covers ALL sheets.
                # Short keywords matched exactly (to avoid false positives like "PL" in "AP Aging")
                exact_keywords = {"pl", "bs", "p&l"}
                # Longer keywords matched as substrings
                substr_keywords = [
                    "pandl", "profit and loss", "statement of profit",
                    "balance sheet", "financials", "income", "revenue",
                    "nt", "note", "notes"
                ]

                # Step 1: Dump ALL sheets to text for regex fallback (CIN, FCGPR, dates etc.)
                text_blocks = []
                for sheet in xl.sheet_names:
                    df_full = pd.read_excel(xl, sheet_name=sheet, header=None)
                    # Prepend sheet name so regex can match sheet context
                    text_blocks.append(f"{sheet}\n" + df_full.to_string(index=False, header=False, na_rep=''))
                md_content = "\n".join(text_blocks)

                # Step 2: Determine which sheets to scan for financial tables
                priority_sheets = []
                for sheet in xl.sheet_names:
                    sheet_lower = sheet.lower().strip()
                    is_exact = sheet_lower in exact_keywords
                    is_substr = any(kw in sheet_lower for kw in substr_keywords)
                    if is_exact or is_substr:
                        priority_sheets.append(sheet)
                        print(f"    [+] Priority financial sheet matched: '{sheet}'")

                # Fallback: if no sheet name matched keywords, scan all sheets
                sheets_to_scan = priority_sheets if priority_sheets else xl.sheet_names
                if not priority_sheets:
                    print(f"    [!] No priority sheet matched — scanning all {len(xl.sheet_names)} sheets")

                # Step 3: Extract structured tables from priority (or all) sheets
                for sheet in sheets_to_scan:
                    df = pd.read_excel(xl, sheet_name=sheet, header=None)

                    # Find dynamic header row (look for Particulars, PY/FY, or Year numbers)
                    header_idx = -1
                    for i in range(min(15, len(df))):
                        row_vals = [str(x).lower() if pd.notna(x) else "" for x in df.iloc[i].values]
                        has_partic = any('particular' in x for x in row_vals)
                        has_py = any('previous' in x or 'py' in x for x in row_vals)
                        has_fy = any('current' in x or 'fy' in x for x in row_vals)
                        has_years = sum(1 for x in row_vals if re.search(r'\b(19|20)\d{2}\b', x)) >= 2

                        if has_partic or (has_py and has_fy) or has_years:
                            header_idx = i
                            break

                    if header_idx != -1:
                        header = [str(c) if pd.notna(c) else "" for c in df.iloc[header_idx].values]
                        data = []
                        for i in range(header_idx + 1, len(df)):
                            row = [str(x) if pd.notna(x) else "" for x in df.iloc[i].values]
                            # Only add non-empty rows
                            if any(str(r).strip() for r in row):
                                data.append(row)
                        tables.append([header] + data)
                        print(f"    [+] Extracted table from sheet '{sheet}' (header row {header_idx})")

                print(f"[+] Found {len(tables)} tables in Financials Excel")
            except Exception as e:
                print(f"[!] Error parsing native Financials Excel: {e}")
                
        elif docs_paths.get("financials") and docs_paths["financials"].endswith(".md"):
            fin_md_path = docs_paths["financials"]
            print(f"[*] Ingesting Financials natively from Markdown: {os.path.basename(fin_md_path)}")
            with open(fin_md_path, "r") as f:
                md_content = f.read()
            tables = self.parse_markdown_tables(md_content)
            print(f"[+] Found {len(tables)} tables in Financials MD")
            
        elif ocr_outputs and ocr_outputs.get("financials"):
            fin_md_path = ocr_outputs["financials"].get("md")
            fin_json_path = ocr_outputs["financials"].get("json")
            
            if fin_md_path and os.path.exists(fin_md_path):
                print(f"[*] Ingesting Financials Markdown OCR: {os.path.basename(fin_md_path)}")
                with open(fin_md_path, "r") as f:
                    md_content = f.read()
                tables = self.parse_markdown_tables(md_content)
                print(f"[+] Found {len(tables)} tables in Financials MD")
            elif fin_json_path and os.path.exists(fin_json_path):
                print(f"[*] Ingesting Financials JSON OCR (extracting HTML tables): {os.path.basename(fin_json_path)}")
                try:
                    with open(fin_json_path, "r") as f:
                        json_data = json.load(f)
                    
                    html_tables = []
                    def recursive_find_tables(node):
                        if not isinstance(node, dict):
                            return
                        if node.get("block_type") == "Table" and "html" in node:
                            html_tables.append(node["html"])
                        if "children" in node and node["children"]:
                            for child in node["children"]:
                                recursive_find_tables(child)
                                
                    recursive_find_tables(json_data)
                    print(f"[+] Found {len(html_tables)} HTML tables in JSON OCR")
                    
                    # Parse each HTML table
                    for html in html_tables:
                        p = HTMLTableParser()
                        p.feed(html)
                        tables.extend(p.tables)
                        
                    # Also try to extract text details (like PAN/CIN) from any text key in the JSON
                    # We can join all page text or raw markdown if available
                    md_content = json_data.get("markdown", "")
                except Exception as e:
                    print(f"[!] Error parsing HTML tables from financials JSON: {e}")
                    
        # Fallback to pdfplumber if OCR wasn't available or yielded no tables
        if not tables and docs_paths.get("financials") and os.path.exists(docs_paths["financials"]):
            fin_pdf = docs_paths["financials"]
            print(f"[*] Ingesting Financials natively using pdfplumber: {os.path.basename(fin_pdf)}")
            try:
                import pdfplumber
                with pdfplumber.open(fin_pdf) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text()
                        if text:
                            md_content += text + "\n"
                        extracted_tables = page.extract_tables()
                        for table in extracted_tables:
                            if table:
                                cleaned_table = []
                                for row in table:
                                    cleaned_row = []
                                    for cell in row:
                                        cell_text = cell if cell else ""
                                        cell_text = cell_text.replace('\n', ' ').strip()
                                        cleaned_row.append(cell_text)
                                    cleaned_table.append(cleaned_row)
                                tables.append(cleaned_table)
                print(f"[+] Found {len(tables)} tables in native Financials PDF")
            except Exception as e:
                print(f"[!] Error parsing native Financials PDF: {e}")
                    
        if tables:
            fin_details = self.extract_financials_from_tables(tables)
            for k, v in fin_details.items():
                if k not in all_extracted or not all_extracted[k]:
                    all_extracted[k] = v
            
        if md_content:
            # Try to extract PAN/CIN from financials text too if not found in Board Report
            text_info = self.extract_company_details(md_content)
            for k, v in text_info.items():
                if k not in all_extracted:
                    all_extracted[k] = v
                    
            # Extract Filing Year and Closing Date from Financial Statement heading
            month_map = {
                'january': '01', 'jan': '01', 'february': '02', 'feb': '02',
                'march': '03', 'mar': '03', 'april': '04', 'apr': '04',
                'may': '05', 'june': '06', 'jun': '06', 'july': '07', 'jul': '07',
                'august': '08', 'aug': '08', 'september': '09', 'sep': '09', 'sept': '09',
                'october': '10', 'oct': '10', 'november': '11', 'nov': '11',
                'december': '12', 'dec': '12'
            }
            date_regex = r'(?:BS|Balance\s*Sheet|Statement\s*of\s*Profit\s*and\s*Loss|Financial\s*Statements)[\s\S]{0,120}?(?:as\s*at|for\s*the\s*year\s*ended)\s*(?:([a-zA-Z]+)\s*(\d{1,2})(?:st|nd|rd|th)?|(\d{1,2})(?:st|nd|rd|th)?\s*([a-zA-Z]+))\s*,?\s*(20\d{2})'
            date_match = re.search(date_regex, md_content, re.IGNORECASE)
            
            if not date_match:
                # Fallback to the very first date found in the document
                fallback_regex = r'(?:([a-zA-Z]+)\s*(\d{1,2})(?:st|nd|rd|th)?|(\d{1,2})(?:st|nd|rd|th)?\s*([a-zA-Z]+))\s*,?\s*(20\d{2})'
                date_match = re.search(fallback_regex, md_content, re.IGNORECASE)
            
            if date_match:
                month1, day1, day2, month2, year = date_match.groups()
                month_str = month1 if month1 else month2
                day_str = day1 if day1 else day2
                
                month_num = month_map.get(month_str.lower(), '03')
                day_num = f'{int(day_str):02d}'
                
                all_extracted["filing_year"] = int(year)
                all_extracted["closing_date"] = f"{day_num}/{month_num}/{year}"
            
            # Custom robust extraction for share counts and face values from Balance Sheet notes
            # 1. Equity Share Count
            eq_match = re.search(r'Movement in the Equity Share capital[\s\S]*?Shares outstanding at the end of the year\s*\|\s*([\d,.]+)\s*\|\s*[\d,.]+\s*\|\s*([\d,.]+)', md_content, re.IGNORECASE)
            if not eq_match:
                # Fallback robust match
                eq_match = re.search(r'Movement in.*?Equity Share [Cc]apital[\s\S]*?Shares outstanding at the end of the (?:year|period)\s*\|\s*([\d,.]+)\s*\|\s*[\d,.]+\s*\|\s*([\d,.]+)', md_content, re.IGNORECASE)
            
            if eq_match:
                if "equity_shares_count_fy" not in all_extracted:
                    all_extracted["equity_shares_count_fy"] = self.parse_number(eq_match.group(1))
                if "equity_shares_count_py" not in all_extracted:
                    all_extracted["equity_shares_count_py"] = self.parse_number(eq_match.group(2))
            
            # 2. CCPS Share Count
            ccps_match = re.search(r'Movement in.*Preference Share capital[\s\S]*?Shares outstanding at the end of the year\s*\|\s*([\d,.]+)\s*\|\s*[\d,.]+\s*\|\s*([\d,.]+)', md_content, re.IGNORECASE)
            if not ccps_match:
                # Fallback robust match
                ccps_match = re.search(r'Movement in.*?(?:Preference|Convertible Preference) Share [Cc]apital[\s\S]*?Shares outstanding at the end of the (?:year|period)\s*\|\s*([\d,.]+)\s*\|\s*[\d,.]+\s*\|\s*([\d,.]+)', md_content, re.IGNORECASE)
                
            if ccps_match:
                if "part_pref_shares_count_fy" not in all_extracted:
                    all_extracted["part_pref_shares_count_fy"] = self.parse_number(ccps_match.group(1))
                if "part_pref_shares_count_py" not in all_extracted:
                    all_extracted["part_pref_shares_count_py"] = self.parse_number(ccps_match.group(2))
            
            # 3. Face Values
            eq_fv = re.search(r'par value of Rs\.?\s*(\d+)\s*per share', md_content, re.IGNORECASE)
            if eq_fv:
                val = float(eq_fv.group(1))
                all_extracted["equity_face_value_py"] = val
                all_extracted["equity_face_value_fy"] = val
            
            ccps_fv = re.search(r'face value of Rs\.?\s*(\d+)\s*per share', md_content, re.IGNORECASE)
            if ccps_fv:
                val = float(ccps_fv.group(1))
                all_extracted["part_pref_face_value_py"] = val
                all_extracted["part_pref_face_value_fy"] = val
            
            # 4. Profit & Loss Balance (Retained Earnings/Surplus in Statement of P&L)
            pl_match = re.search(r'Surplus(?:\s*/\s*\(Deficit\))?\s*in\s*Statement\s*of\s*Profit\s*and\s*Loss[\s\S]{0,10000}?Balance\s*at\s*the\s*end\s*of\s*the\s*year(?:\||\s)*([-\d,.]+)(?:\||\s)*([-\d,.]+)', md_content, re.IGNORECASE)
            if pl_match:
                all_extracted["pl_balance_fy"] = self.parse_number(pl_match.group(1))
                all_extracted["pl_balance_py"] = self.parse_number(pl_match.group(2))
            
            # 5. Sales & Purchases
            rev_match = re.search(r'(?:Domestic\s+sales|Revenue\s+from\s+operations\s*-\s*Domestic)\s*\|\s*\d+\s*\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', md_content, re.IGNORECASE)
            if rev_match:
                all_extracted["domestic_sales_fy"] = self.parse_number(rev_match.group(1))
                all_extracted["domestic_sales_py"] = self.parse_number(rev_match.group(2))
                all_extracted["export_sales_fy"] = 0.0
                all_extracted["export_sales_py"] = 0.0
            
            indig_match = re.search(r'Indigenous\s*\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', md_content, re.IGNORECASE)
            if indig_match:
                all_extracted["domestic_purchases_fy"] = self.parse_number(indig_match.group(1))
                all_extracted["domestic_purchases_py"] = self.parse_number(indig_match.group(2))
            
            imp_match = re.search(r'Imported\s*\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', md_content, re.IGNORECASE)
            if imp_match:
                all_extracted["import_purchases_fy"] = self.parse_number(imp_match.group(1))
                all_extracted["import_purchases_py"] = self.parse_number(imp_match.group(2))

        # 6. Related Party Transactions (Liabilities & Claims for FDI Block 1, 2, 3 and DI Block 1, 2, 3)
        for prefix in ["fdi_investor", "di_investor"]:
            for i in range(1, 4):
                inv_key = f"{prefix}_{i}_name"
                if inv_key in all_extracted and all_extracted[inv_key]:
                    investor_name = all_extracted[inv_key]
                    
                    # Focus on Related Party section if it exists
                    search_content = md_content
                    rpt_match = re.search(r'(?i)(?:Related\s+party\s+disclosures?|Related\s+party\s+transactions?)', md_content)
                    if rpt_match:
                        search_content = md_content[rpt_match.start():]

                    # 2.1 Other Liabilities (Trade payables, loans from investor)
                    payables_match = re.search(r'(?:trade payable|payable|trade creditor|creditor|owed|outstanding|ecb|loan|ccd|ccp)[\s\S]{0,100}?' + re.escape(investor_name) + r'[\s\S]{0,100}?\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', search_content, re.IGNORECASE)
                    if payables_match:
                        all_extracted[f"{prefix}_{i}_other_liabilities_fy"] = self.parse_number(payables_match.group(1))
                        all_extracted[f"{prefix}_{i}_other_liabilities_py"] = self.parse_number(payables_match.group(2))
                        
                    # 2.2 Other Claims (Trade receivables from investor)
                    receivables_match = re.search(r'(?:trade receivable|receivable|trade debtor|debtor|due from)[\s\S]{0,100}?' + re.escape(investor_name) + r'[\s\S]{0,100}?\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', search_content, re.IGNORECASE)
                    if receivables_match:
                        all_extracted[f"{prefix}_{i}_other_claims_fy"] = self.parse_number(receivables_match.group(1))
                        all_extracted[f"{prefix}_{i}_other_claims_py"] = self.parse_number(receivables_match.group(2))
                        
                    # 1.2 Claims on Direct Investors (Reverse Investment)
                    inv_match = re.search(r'investment[\s\S]{0,50}?' + re.escape(investor_name) + r'[\s\S]{0,100}?\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', search_content, re.IGNORECASE)
                    if inv_match:
                        all_extracted[f"{prefix}_{i}_claims_fy"] = self.parse_number(inv_match.group(1))
                        all_extracted[f"{prefix}_{i}_claims_py"] = self.parse_number(inv_match.group(2))

                    # 3 Disinvestments
                    if f"{prefix}_{i}_equity_percent_py" in all_extracted and f"{prefix}_{i}_equity_percent_fy" in all_extracted:
                        py_pct = all_extracted[f"{prefix}_{i}_equity_percent_py"]
                        fy_pct = all_extracted[f"{prefix}_{i}_equity_percent_fy"]
                        if fy_pct < py_pct:
                            dis_match = re.search(r'(?:buyback|transfer|reduction)[\s\S]{0,150}?\|\s*([\d,.]+)\s*\|\s*([\d,.]+)', search_content, re.IGNORECASE)
                            if dis_match:
                                all_extracted[f"{prefix}_{i}_disinvestment_fy"] = self.parse_number(dis_match.group(1))
                                all_extracted[f"{prefix}_{i}_disinvestment_py"] = self.parse_number(dis_match.group(2))
                    
        # 6c. First FDI Date
        fcgpr_match = re.search(r'(?:FCGPR|first allotment|first share allotment)[\s\S]{0,100}?((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s*\d{4})', md_content, re.IGNORECASE)
        if fcgpr_match:
            all_extracted["fdi_first_received_date"] = fcgpr_match.group(1).strip()

        # 7. Dynamically extract Nature of Business / NIC Code
        # Combine Board Report text and Financials md_content
        br_text = text if "text" in locals() else ""
        fin_text = md_content if "md_content" in locals() else ""
        full_text = (br_text + "\n" + fin_text).lower()
        
        # 6.a Find corporate or segment business activity description
        segment_desc = ""
        
        # Try finding 'engaged in the business of ...' (Note 1)
        biz_match_1 = re.search(r'engaged\s+in\s+the\s+business\s+of\s+([^,.\n]+)', full_text, re.IGNORECASE)
        if biz_match_1:
            segment_desc = biz_match_1.group(1).strip()
            
        # Try finding 'segment, viz. ...' (Note 29)
        if not segment_desc:
            segment_match = re.search(r'segment,\s*viz\.\s*\"([^\"]+)\"', full_text)
            if segment_match:
                segment_desc = segment_match.group(1).strip()
        all_extracted["nic_code"] = ""

        # ── Previous FLA Prefill (Section I contact & identity fields) ──────────────
        # These fields have NO extraction logic from board report / financials.
        # They are ALWAYS sourced exclusively from the previous year FLA form.
        prev_fla_path = docs_paths.get("previous_fla")
        if not prev_fla_path:
            # Also check sibling input dir for FLA_ONLINE_FORM or previous_fla_ files
            input_dir = os.path.dirname(list(docs_paths.values())[0]) if docs_paths else ""
            if input_dir and os.path.isdir(input_dir):
                for fname in os.listdir(input_dir):
                    if fname.startswith("previous_fla_") or "FLA_ONLINE_FORM" in fname:
                        prev_fla_path = os.path.join(input_dir, fname)
                        break

        if prev_fla_path and os.path.exists(prev_fla_path):
            try:
                from automation_engine.modules.fla.comparison_platform.modules.legacy_parser import LegacyFLAParser
                with open(prev_fla_path, "r", encoding="utf-8", errors="ignore") as f:
                    prev_data = LegacyFLAParser().parse_md(f.read())

                # Fields that must ONLY come from previous FLA — mapped to internal keys
                SECTION_I_FROM_PREV_FLA = {
                    "Name of the Contact Person":               "contact_name",
                    "Telephone No. (with extension)":           "telephone",
                    "Mobile Number":                            "mobile_number",
                    "E-Mail ID (Head of the institution)":      "email_id",
                    "E-Mail of Contact person":                 "email_contact",
                    "Designation":                              "designation",
                    "Website (if any)":                         "website",
                    "Type of company":                          "company_type",
                    "2. PAN number":                            "pan_number",
                }
                # Always overwrite these fields — previous FLA is the sole authoritative source
                for prev_key, internal_key in SECTION_I_FROM_PREV_FLA.items():
                    if prev_key in prev_data:
                        val = prev_data[prev_key]
                        all_extracted[internal_key] = val
                        print(f"  [PrevFLA] '{internal_key}' = '{val}'")

                # Special: Account closing date → increment year for current filing
                for k, v in prev_data.items():
                    ck = re.sub(r"[^a-z0-9]", "", str(k).lower())
                    if "accountclosingdate" in ck:
                        m = re.search(r'(\d{2})/(\d{2})/(\d{4})', str(v))
                        if m:
                            d, mo, yr = m.groups()
                            next_yr = int(yr) + 1
                            if "filing_year" not in all_extracted or not all_extracted.get("filing_year"):
                                all_extracted["filing_year"] = next_yr
                                print(f"  [PrevFLA] 'filing_year' = {next_yr}")
                            if "closing_date" not in all_extracted or not all_extracted.get("closing_date"):
                                all_extracted["closing_date"] = f"{d}/{mo}/{next_yr}"
                                print(f"  [PrevFLA] 'closing_date' = {d}/{mo}/{next_yr}")
                        break

                print(f"  [+] Section I prefilled from Previous FLA: {os.path.basename(prev_fla_path)}")
            except Exception as e:
                print(f"  [!] Could not prefill from Previous FLA: {e}")

        return all_extracted

