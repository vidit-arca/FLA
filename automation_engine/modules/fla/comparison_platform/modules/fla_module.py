import pandas as pd
import openpyxl
import re
from typing import Dict, Any, List
from ..base import BaseComparisonPlatformModule


class FLAComparisonModule(BaseComparisonPlatformModule):
    def __init__(self, template_path: str = "/Users/apple/Desktop/FLA/automation_engine/modules/fla/excel/FLA_comparsion.xlsx"):
        self.template_path = template_path

    def _safe_write(self, ws, row, col, value):
        """Write a value to a cell, safely skipping merged cells."""
        try:
            cell = ws.cell(row=row, column=col)
            cell.value = value
            return True
        except AttributeError:
            # MergedCell — skip it
            return False

    def _normalize_val(self, val: Any) -> str:
        if pd.isna(val) or val is None:
            return ""
        s = str(val).strip()
        if s.lower() in ["nan", "nat", "none", "null"]:
            return ""
        return s

    def _clean_key(self, s: str) -> str:
        """Normalise a field name for fuzzy matching."""
        s = re.sub(r'[^a-z0-9 ]', '', s.lower())
        return re.sub(r'\s+', ' ', s).strip()

    def _lookup(self, parsed: Dict[str, Any], field_name: str) -> Any:
        """
        Look up a field in the parsed dictionary.
        Tries exact match first, then normalised fuzzy match.
        """
        # 1. Exact match
        if field_name in parsed:
            return parsed[field_name]

        # 2. Normalised match
        clean = self._clean_key(field_name)
        for k, v in parsed.items():
            if self._clean_key(k) == clean:
                return v

        if len(clean) >= 10 and len(clean) <= 100:
            for k, v in parsed.items():
                ck = self._clean_key(k)
                if clean in ck:
                    return v

        return None

    def compare(self, source_path: str, target_path: str) -> List[Dict[str, Any]]:
        results = []

        try:
            # Load the comparison template (rules)
            template_dfs = pd.read_excel(self.template_path, sheet_name=None, header=None)

            from .legacy_parser import LegacyFLAParser
            legacy_parser = LegacyFLAParser()

            # Handle source_path: Excel vs Markdown
            if source_path.endswith('.md'):
                with open(source_path, 'r', encoding='utf-8') as f:
                    md_text = f.read()
                parsed_previous_fla = legacy_parser.parse_md(md_text)
            else:
                source_dfs = pd.read_excel(source_path, sheet_name=None, engine="openpyxl")
                parsed_previous_fla = legacy_parser.parse_previous_fla(source_dfs)

            # target_path is Current FLA (skeletal or partially filled)
            target_dfs = pd.read_excel(target_path, sheet_name=None, header=None, engine="openpyxl")
            target_wb = openpyxl.load_workbook(target_path)
            has_modifications = False

        except Exception as e:
            raise Exception(f"Failed to load excel files: {str(e)}")

        for sheet_name, template_df in template_dfs.items():
            target_ws = target_wb[sheet_name] if sheet_name in target_wb.sheetnames else None
            if not target_ws or sheet_name not in target_dfs:
                continue
            

            tgt_df = target_dfs[sheet_name]

            # ── Dynamic column detection ──
            field_name_col = -1
            particulars_col = -1
            rule_col = -1
            py_share_col = -1   # "Shares in Actual - End March PY"
            py_amount_col = -1  # "Amount in Rs. LAKHS - End March PY"
            fy_share_col = -1   # "Shares in Actual - End March FY"
            fy_amount_col = -1  # "Amount in Rs. LAKHS - End March FY"
            head_idx = 0

            for hi in range(min(10, len(template_df))):
                for c_idx, val in template_df.iloc[hi].items():
                    s = str(val).strip().lower()
                    if s == "field name":
                        field_name_col = c_idx
                    elif s == "particulars":
                        particulars_col = c_idx
                    elif s == "source of information":
                        rule_col = c_idx
                    elif "end march py" in s:
                        if py_share_col == -1:
                            py_share_col = c_idx
                        else:
                            py_amount_col = c_idx
                    elif "end march fy" in s:
                        if fy_share_col == -1:
                            fy_share_col = c_idx
                        else:
                            fy_amount_col = c_idx

                if field_name_col != -1:
                    head_idx = hi

            # Fallback: if we found 'Source of Information' but NOT 'Field Name',
            # assume field names are in column 1 (e.g. Section IV layout)
            if field_name_col == -1 and rule_col != -1:
                field_name_col = 1
                head_idx = 0  # No header row found, start from row 0

            # ── Route A: Sheets with explicit "Source of Information" column ──
            if field_name_col != -1 and rule_col != -1:
                for idx, row in template_df.iterrows():
                    if idx <= head_idx:
                        continue

                    field_name = self._normalize_val(row.iloc[field_name_col])
                    if not field_name:
                        continue

                    # Strict Scope Filter: Only Section II (blocks 3, 4, 5, 6) and Section III
                    if sheet_name not in ['Section II', 'Section III']:
                        continue
                    if sheet_name == 'Section II' and idx < 24:
                        continue

                    rule_text = self._normalize_val(row.iloc[rule_col]).upper()

                    if "PREVIOUS FLA" in rule_text:
                        target_val_raw = None
                        write_col = particulars_col if particulars_col != -1 else 3

                        if idx < len(tgt_df) and write_col < len(tgt_df.columns):
                            target_val_raw = tgt_df.iloc[idx, write_col]

                        source_val_raw = self._lookup(parsed_previous_fla, field_name)

                        prev_val = self._normalize_val(source_val_raw)
                        curr_val = self._normalize_val(target_val_raw)

                        final_val, reason = self._apply_rule(rule_text, prev_val, curr_val)

                        if final_val != curr_val and final_val != "":
                            row_idx = idx + 1
                            col_idx = write_col + 1
                            if self._safe_write(target_ws, row_idx, col_idx, final_val):
                                has_modifications = True

                        results.append({
                            "fieldName": f"{sheet_name} | {field_name}",
                            "mappingType": rule_text,
                            "previousValue": prev_val,
                            "currentValue": curr_val,
                            "finalSelectedValue": final_val,
                            "reason": reason
                        })

            # ── Route B: Sheets with PY/FY columns (Section II, III, IV) ──
            # Fill BOTH PY and FY columns from the previous FLA data.
            # Dynamically track PY/FY headers as they can appear in repeated blocks (e.g. Sect III, IV)
            if field_name_col != -1:
                curr_py_share = py_share_col
                curr_py_amount = py_amount_col
                curr_fy_share = fy_share_col
                curr_fy_amount = fy_amount_col

                for idx, row in template_df.iterrows():
                    # Detect PY/FY headers in the current row
                    row_vals = [str(v).strip().lower() for v in row.values]
                    if "end march py" in row_vals or "end march fy" in row_vals:
                        curr_py_share = -1
                        curr_py_amount = -1
                        curr_fy_share = -1
                        curr_fy_amount = -1
                        for c_idx, val in row.items():
                            s = str(val).strip().lower()
                            if "end march py" in s:
                                if curr_py_share == -1: curr_py_share = c_idx
                                else: curr_py_amount = c_idx
                            elif "end march fy" in s:
                                if curr_fy_share == -1: curr_fy_share = c_idx
                                else: curr_fy_amount = c_idx
                        continue

                    if curr_py_share == -1 and curr_fy_share == -1:
                        continue  # No active PY/FY block

                    if idx <= head_idx:
                        continue

                    field_name = self._normalize_val(row.iloc[field_name_col])
                    if not field_name:
                        field_name = self._normalize_val(row.iloc[0])
                        if not field_name and len(row) > 2:
                            # Try column 3 as a fallback (sometimes used for sub-items)
                            field_name = self._normalize_val(row.iloc[2])
                    if not field_name:
                        continue

                    # Filter out purely structural rows
                    fn_lower = field_name.lower()
                    if 'auto-calculated' in fn_lower or 'total' in fn_lower:
                        continue

                    # FY value = the main key (previous FLA's "End March 2025" data)
                    fy_val = self._normalize_val(self._lookup(parsed_previous_fla, field_name))
                    # PY value = the __PY key (previous FLA's "End March 2024" data)
                    py_val = self._normalize_val(parsed_previous_fla.get(f"{field_name}__PY", None))
                    # If no separate PY, use FY as fallback
                    if not py_val:
                        py_val = fy_val

                    # Disable share columns and FY Amount column for Financial rows (Row 26+) in Section II
                    # These rows only take the PY Amount from the legacy data.
                    tmp_py_share = curr_py_share
                    tmp_fy_share = curr_fy_share
                    tmp_fy_amount = curr_fy_amount
                    if sheet_name == 'Section II' and idx >= 25:
                        tmp_py_share = -1
                        tmp_fy_share = -1
                        tmp_fy_amount = -1

                    # Fill all 4 column slots: PY Shares, PY Amount, FY Shares, FY Amount
                    col_map = [
                        (tmp_py_share, py_val, "PY Shares/Col1", True),
                        (curr_py_amount, py_val, "PY Amount/Col2", False),
                        (tmp_fy_share, fy_val, "FY Shares/Col1", True),
                        (tmp_fy_amount, fy_val, "FY Amount/Col2", False),
                    ]

                    # Scope Filter: Only validate Section II (blocks 3+) and Section III
                    if sheet_name not in ['Section II', 'Section III']:
                        continue
                    if sheet_name == 'Section II' and idx < 24:
                        continue


                    for col_idx, fill_val, label, is_share in col_map:
                        if col_idx == -1 or idx >= len(tgt_df):
                            continue
                            
                        # Exclude Shareholding data from validation
                        if is_share:
                            continue


                        curr = self._normalize_val(
                            tgt_df.iloc[idx, col_idx] if col_idx < len(tgt_df.columns) else None
                        )
                        
                        if sheet_name == 'Section II' and idx == 25:
                            print(f"[DEBUG IN MODULE] label={label}, col_idx={col_idx}, curr={curr}, fill_val={fill_val}")

                        if not curr and not fill_val:
                            # Both missing, no issue
                            continue
                            
                        if curr and not fill_val:
                            results.append({
                                "fieldName": f"{sheet_name} | {field_name} ({label})",
                                "mappingType": "STRICT_VALIDATION",
                                "previousValue": "Missing",
                                "currentValue": curr,
                                "finalSelectedValue": curr,
                                "reason": "Mismatch Detected -> Needs Review"
                            })
                        elif not curr and fill_val:
                            results.append({
                                "fieldName": f"{sheet_name} | {field_name} ({label})",
                                "mappingType": "STRICT_VALIDATION",
                                "previousValue": fill_val,
                                "currentValue": "Missing",
                                "finalSelectedValue": "",
                                "reason": "Missing in Current -> Needs Review"
                            })
                        elif curr == fill_val:
                            results.append({
                                "fieldName": f"{sheet_name} | {field_name} ({label})",
                                "mappingType": "STRICT_VALIDATION",
                                "previousValue": fill_val,
                                "currentValue": curr,
                                "finalSelectedValue": curr,
                                "reason": "Values Match Perfectly"
                            })
                        else:
                            results.append({
                                "fieldName": f"{sheet_name} | {field_name} ({label})",
                                "mappingType": "STRICT_VALIDATION",
                                "previousValue": fill_val,
                                "currentValue": curr,
                                "finalSelectedValue": curr,
                                "reason": "Mismatch Detected -> Needs Review"
                            })

        # Save
        if has_modifications:
            target_wb.save(target_path)

        return results

    def _apply_rule(self, rule_text: str, prev_val: str, curr_val: str):
        """Apply the PREVIOUS FLA / UNLESS rule logic."""
        if "UNLESS" in rule_text:
            if prev_val and not curr_val:
                return prev_val, "Current Value Missing -> Used Previous"
            elif prev_val and curr_val and prev_val == curr_val:
                return prev_val, "Current Value Same -> Used Previous"
            elif prev_val and curr_val and prev_val != curr_val:
                return curr_val, "Current Value Different -> Used Current"
            elif not prev_val and curr_val:
                return curr_val, "Previous Missing -> Used Current"
            else:
                return "", "Both Missing -> Left Blank"
        else:
            if curr_val:
                return curr_val, "Current Value Present -> Used Current"
            else:
                return prev_val, "Current Value Missing -> Used Previous"
