import json
import os
import re
import openpyxl
from automation_engine.modules.aoc4.aoc4_error_checker import AOC4CommonErrorEngine
from automation_engine.modules.aoc4.compliance_engine import PrivateComplianceEngine
from automation_engine.modules.aoc4.rpt_loans_engine import RPTLoansEngine
from automation_engine.modules.aoc4.excel_extractor import AOC4ExcelExtractor

class AOC4RuleEngine:
    def __init__(self, config_path: str):
        self.config_path = config_path
        with open(config_path, "r") as f:
            self.config = json.load(f)
            
        base_dir = os.path.dirname(config_path)
        self.excel_path = os.path.join(base_dir, "excel", "ANNFIL COMMONERROR .xlsx")
        self.output_skeletal_path = os.path.join(os.path.dirname(__file__), "excel", "ANNFIL COMMONERROR .xlsx")
        self.checker = AOC4CommonErrorEngine(self.excel_path)
        self.compliance_engine = PrivateComplianceEngine()
        self.rpt_loans_engine = RPTLoansEngine()
        self.excel_extractor = AOC4ExcelExtractor()

    def _normalize_string(self, text):
        if not text:
            return ""
        return str(text).strip().lower().replace(" ", "").replace("\n", "")

    def evaluate_all(self, extracted_data: dict) -> dict:
        print("[*] AOC 4 Rule Engine: Evaluating common errors...")
        

        print("[*] AOC 4 Rule Engine: Evaluating private compliance...")
        # 1. Run Excel Extractor on the original docs to pull structured financial metrics
        docs = extracted_data.get("docs", {})
        financial_data = self.excel_extractor.extract_from_docs(docs)
        
        # Identify missing numeric keys
        expected_keys = list(self.excel_extractor.numeric_keywords.keys())
        missing_keys = [k for k in expected_keys if financial_data.get(k) is None]
        
        # If any numeric fields are missing, try extracting from raw Markdown (PDFs)
        if missing_keys and "full_text" in extracted_data:
            from automation_engine.modules.aoc4.parser import AOC4Parser
            # The parser logic is now in AOC4Parser. We need an instance.
            # But wait, self doesn't have an instance of AOC4Parser readily available here, 
            # so we instantiate one temporarily or call it statically.
            temp_parser = AOC4Parser(self.config_path)
            fallback_data = temp_parser.extract_financials_from_text(extracted_data["full_text"], missing_keys)
            
            scale = self.excel_extractor.detect_financials_scale(extracted_data["full_text"])
            if scale != 1.0:
                print(f"[*] AOC 4 Parser: Applying unit scale multiplier {scale} to text fallback data")
                for k in fallback_data.keys():
                    if isinstance(fallback_data[k], (int, float)):
                        fallback_data[k] = fallback_data[k] * scale
            
            # Merge the fallback data
            for k, v in fallback_data.items():
                financial_data[k] = v
                
        # Fallback for Schedule III format detection from Markdown text
        if financial_data.get("has_schedule_iii_format") is None and "full_text" in extracted_data:
            text_lower = extracted_data["full_text"].lower()
            required_headers = {
                "equity and liabilities", "shareholders' funds", "non-current liabilities",
                "current liabilities", "assets", "non-current assets", "current assets"
            }
            found_count = sum(1 for req in required_headers if req in text_lower or req.replace("'", "") in text_lower)
            if found_count >= 5:
                financial_data["has_schedule_iii_format"] = "yes"
                
        # Merge all processed financial metrics back into the main extracted_data for Excel mapping
        for k, v in financial_data.items():
            if v is not None or k not in extracted_data:
                extracted_data[k] = v
        
        # 1. Run the Compliance Engine with the populated numerical data
        compliance_flags = self.compliance_engine.execute(extracted_data)
        
        # 2. Run common error checker so it has access to the updated extracted_data (e.g. is_small_company)
        common_flags = self.checker.execute(extracted_data)
        
        # 3. Run the RPT & Loans Engine
        print("[*] AOC 4 Rule Engine: Evaluating RPT and Loans...")
        rpt_flags = self.rpt_loans_engine.execute(extracted_data)
        
        flags = common_flags + compliance_flags + rpt_flags
        
        target_cells = {
            "Common Error": {},
            "compliance for Private ": {},
            "RPT and loans to Director": {}
        }
        
        # Open the target skeletal template to find dynamic rows
        if not os.path.exists(self.output_skeletal_path):
            print(f"[!] Warning: Output skeletal path not found at {self.output_skeletal_path}")
            extracted_data["flags"] = flags
            return target_cells
            
        wb = openpyxl.load_workbook(self.output_skeletal_path, data_only=True)
        if "Common Error" in wb.sheetnames:
            sheet = wb["Common Error"]
            # Build a map of Particulars -> Row Number
            row_map = {}
            for row in range(2, sheet.max_row + 1):
                particulars = sheet.cell(row=row, column=2).value
                if particulars:
                    norm = self._normalize_string(particulars)
                    row_map[norm] = row
                    
            for flag in common_flags:
                norm_flag = self._normalize_string(flag["particulars"])
                matched_row = row_map.get(norm_flag)
                if matched_row:
                    yes_no = flag.get("user_value") if flag.get("user_value") else ("No" if flag.get("status") == "Failed" else "Yes")
                    comment = flag.get("reason", "")
                    
                    target_cells["Common Error"][f"C{matched_row}"] = yes_no
                    target_cells["Common Error"][f"D{matched_row}"] = comment
                else:
                    print(f"[!] Could not find dynamic mapping for rule in Excel: {flag['particulars'][:50]}...")
                    
        # We can also map compliance flags to 'compliance for Private ' if needed
        if "compliance for Private " in wb.sheetnames:
            sheet_comp = wb["compliance for Private "]
            row_map_comp = {}
            for row in range(2, sheet_comp.max_row + 1):
                # Requirement is in Column C (3)
                particulars = sheet_comp.cell(row=row, column=3).value
                if particulars:
                    norm = self._normalize_string(particulars)
                    row_map_comp[norm] = row
            
            raw_data_map = {
                "Paidup capital": ("paid_up_capital", "prev_paid_up_capital"),
                "Reserves and Surplus": ("reserves_and_surplus", None),
                "Total Borrowings": ("borrowings", None),
                "Loangiven by Company to Directors or Director related entities (assets)": ("loan_to_directors_assets", None),
                "Loans given by Company (assets)": ("loan_given_by_company", None),
                "Investments made by Company (assets)": ("investments_made", None),
                "Corporate Guarantees given by Company": ("corporate_guarantees", None),
                "Loan from Directors or their relatives (Liabilities)": ("loan_from_directors", None),
                "Secured Loan": ("secured_loan", None),
                "Advance from Customers, Shareholders, Security Deposits (Liabilitys)": ("advance_from_customers", None),
                "Dues to MSME": ("dues_to_msme", None),
                "Networth": ("net_worth", "prev_net_worth"),
                "Turnover": ("turnover", "prev_turnover"),
                "Total Revenue": ("total_revenue", None),
                "Profit Before Tax": ("net_profit_before_tax", None),
                "ED / WTD - 1 Monthly Remuneration": ("rpt_monthly_remun", None),
                "ED / WTD - 2 Monthly Remuneration": ("rpt_monthly_remun_2", None),
                "Number of Bodies Corporate Shareholder holding more than 10%": ("has_corporate_shareholders", None),
                "Is the Company a Holding or a Subsidiary Company ": ("is_subsidiary_or_holding", None),
                "Is the Company a Holding, Subsidiary or Associate of IND AS applicable Companies": ("is_ind_as", None),
                "Exprt": ("export_sales", "prev_export_sales"),
                "Sitting Fees to Directors": ("sitting_fees", "prev_sitting_fees")
            }
            
            if extracted_data.get("paid_up_capital") is not None and extracted_data.get("reserves_and_surplus") is not None:
                try:
                    extracted_data["net_worth"] = float(extracted_data["paid_up_capital"]) + float(extracted_data["reserves_and_surplus"])
                except (ValueError, TypeError):
                    pass
            if extracted_data.get("prev_paid_up_capital") is not None and extracted_data.get("prev_reserves_and_surplus") is not None:
                try:
                    extracted_data["prev_net_worth"] = float(extracted_data["prev_paid_up_capital"]) + float(extracted_data["prev_reserves_and_surplus"])
                except (ValueError, TypeError):
                    pass

            for particular_name, (cy_key, py_key) in raw_data_map.items():
                norm = self._normalize_string(particular_name)
                matched_row = row_map_comp.get(norm)
                if matched_row:
                    if cy_key:
                        val = extracted_data.get(cy_key)
                        if val is None:
                            if cy_key in ["is_subsidiary_or_holding", "is_ind_as"]:
                                val = "No"
                            else:
                                val = 0

                        if cy_key == "has_corporate_shareholders":
                            val = 1 if str(val).lower() == "yes" else (val if isinstance(val, (int, float)) else 0)
                        elif isinstance(val, bool):
                            if not val and cy_key == "is_subsidiary_or_holding":
                                val = "Not a Subsidiary company"
                            else:
                                val = "Yes" if val else "No"
                        elif str(val).lower() in ["yes", "holding", "subsidiary"]: val = "Yes"
                        elif str(val).lower() == "no":
                            if cy_key == "is_subsidiary_or_holding":
                                val = "Not a Subsidiary company"
                            else:
                                val = "No"
                        target_cells["compliance for Private "][f"D{matched_row}"] = val
                    
                    if py_key and extracted_data.get(py_key) is not None:
                        val_py = extracted_data[py_key]
                        target_cells["compliance for Private "][f"F{matched_row}"] = val_py
                        
            # Dynamically map the evaluated Python compliance flags into Rows 35-59 by explicit rule ID!
            # This ensures every compliance check lands on the exact matching template row,
            # and clears out any legacy cached ExcelJet links or unused Previous Year formulas.
            id_to_row = {
                "COMP_SMALL_CO": 35,
                "COMP_CARO": 36,
                "COMP_ROTATION": 37,
                "COMP_IND_AS": 38,
                "COMP_XBRL": 39,
                "COMP_VIGIL": 40,
                "COMP_IFC": 41,
                "COMP_INT_AUDIT": 42,
                "COMP_SEC_AUDIT": 43,
                "COMP_KMP": 44,
                "COMP_LOAN_186": 45,
                "COMP_LOAN_DIRECTOR": 46,
                "COMP_COST_AUDIT": 47,
                "COMP_CHARGE_FORM": 48,
                "COMP_AOC_1": 49,
                "COMP_AOC_2": 50,
                "COMP_RPT_OMNIBUS": 51,
                "COMP_CSR": 52,
                "COMP_CSR_COMMITTEE": 53,
                "COMP_DEPOSIT_DEC": 54,
                "COMP_DPT_3": 55,
                "COMP_MSME": 56,
                "COMP_BEN_2": 57,
                "COMP_MGT_8": 58,
                "COMP_MGT_7_CERT": 59
            }
            for flag in compliance_flags:
                row = id_to_row.get(flag.get("id"))
                if row:
                    target_cells["compliance for Private "][f"C{row}"] = flag["particulars"]
                    target_cells["compliance for Private "][f"D{row}"] = flag.get("user_value", "")
                    target_cells["compliance for Private "][f"E{row}"] = flag.get("rationale", "")
                    target_cells["compliance for Private "][f"F{row}"] = ""
                    target_cells["compliance for Private "][f"G{row}"] = ""
        if "RPT and loans to Director" in wb.sheetnames:
            sheet_rpt = wb["RPT and loans to Director"]
            row_map_rpt = {}
            for row in range(2, sheet_rpt.max_row + 1):
                # Search cols 1, 2, and 3 for particulars (some are in C like "Has the Company give loan...")
                particulars = sheet_rpt.cell(row=row, column=1).value or sheet_rpt.cell(row=row, column=2).value or sheet_rpt.cell(row=row, column=3).value
                if particulars:
                    norm = self._normalize_string(particulars)
                    row_map_rpt[norm] = row
            
            for flag in rpt_flags:
                norm_flag = self._normalize_string(flag["particulars"])
                matched_row = row_map_rpt.get(norm_flag)
                if matched_row:
                    # For RPT, write 'actual_value' to Col F
                    if flag.get("actual_value") is not None and str(flag.get("actual_value")) != "0.0":
                        target_cells["RPT and loans to Director"][f"F{matched_row}"] = flag.get("actual_value")
                    
                    # For Section 185 and 186, Applicability goes to Col D. Otherwise Col G.
                    flag_id = flag.get("id", "")
                    if flag_id.startswith("COMP_SEC_185_") or flag_id.startswith("COMP_SEC_186_"):
                        target_cells["RPT and loans to Director"][f"D{matched_row}"] = flag.get("user_value")
                    else:
                        target_cells["RPT and loans to Director"][f"G{matched_row}"] = flag.get("user_value")

        extracted_data["flags"] = flags
        target_cells["_flags"] = flags
        
        return target_cells
