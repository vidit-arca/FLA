import openpyxl

def dump_financials_details(file_path, out_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb["2_FINANCIALS"]
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"=== FULL DETAILS OF SHEET: 2_FINANCIALS ===\n\n")
        for r in range(1, sheet.max_row + 1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
            if any(val is not None for val in row_vals):
                cleaned = []
                for col_idx, v in enumerate(row_vals):
                    if v is not None:
                        col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                        cleaned.append(f"{col_letter}: {repr(v)}")
                f.write(f"Row {r:03d}: " + " | ".join(cleaned) + "\n")
    wb.close()
    print(f"Financials sheet details dumped to {out_path}")

if __name__ == "__main__":
    dump_financials_details(
        "/Users/apple/Desktop/FLA/automation_engine/excel/FLA_Tool_v5_fixed.xlsx",
        "/Users/apple/Desktop/FLA/financials_details.txt"
    )
