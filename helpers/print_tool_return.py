import openpyxl

def print_tool_return_formulas(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb["3_FLA_RETURN"]
    
    print(f"=== {sheet.title} Formulas and Mapping Logic ===")
    for r in range(1, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
        # Check if there is any value in the row
        if any(v is not None for v in row_vals):
            # Print row index and cell values/formulas
            cleaned = []
            for col_idx, v in enumerate(row_vals):
                if v is not None:
                    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                    cleaned.append(f"{col_letter}: {repr(v)}")
            print(f"Row {r:03d}: " + " | ".join(cleaned))
    wb.close()

if __name__ == "__main__":
    print_tool_return_formulas("/Users/apple/Desktop/FLA/automation_engine/excel/FLA_Tool_v5_fixed.xlsx")
