import openpyxl
import pandas as pd
import os

def inspect_file(file_path):
    print(f"=== Inspecting {os.path.basename(file_path)} ===")
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        print(f"Sheets: {wb.sheetnames}")
        for name in wb.sheetnames:
            sheet = wb[name]
            print(f"  Sheet '{name}': {sheet.max_row} rows, {sheet.max_column} columns")
            # Let's read first few rows using pandas for pretty printing
            try:
                df = pd.read_excel(file_path, sheet_name=name, nrows=10)
                print(f"    Columns: {list(df.columns)}")
                print(f"    Preview:\n{df.head(3)}")
            except Exception as e:
                print(f"    Failed to load sheet with pandas: {e}")
        wb.close()
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
    print("\n")

if __name__ == "__main__":
    inspect_file("/Users/apple/Desktop/FLA/excel/FLA_Tool_v5_fixed.xlsx")
    inspect_file("/Users/apple/Desktop/FLA/excel/FLA Return existing skeletal.xlsx")
