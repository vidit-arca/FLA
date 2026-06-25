import openpyxl
import os

def dump_skeletal(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"=== Dumping skeletal structures for {os.path.basename(file_path)} ===")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        for r in range(1, sheet.max_row + 1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
            # filter out empty rows
            if any(val is not None for val in row_vals):
                # Print non-empty rows, truncated if necessary
                cleaned_vals = [str(v)[:40] if v is not None else "" for v in row_vals]
                print(f"Row {r:02d}: {cleaned_vals}")
    wb.close()

if __name__ == "__main__":
    dump_skeletal("/Users/apple/Desktop/FLA/automation_engine/automation_engine/excel/FLA Return existing skeletal.xlsx")
