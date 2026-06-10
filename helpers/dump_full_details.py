import openpyxl

def dump_full_details(file_path, out_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"=== FULL DETAILS OF {file_path} ===\n\n")
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Annex 1':
                continue # Skip large code annex
            sheet = wb[sheet_name]
            f.write(f"\n=================== SHEET: {sheet_name} ===================\n")
            for r in range(1, sheet.max_row + 1):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
                if any(val is not None for val in row_vals):
                    cleaned_vals = [str(v) if v is not None else "" for v in row_vals]
                    f.write(f"Row {r:03d}: {cleaned_vals}\n")
    wb.close()
    print(f"Full details dumped to {out_path}")

if __name__ == "__main__":
    dump_full_details(
        "/Users/apple/Desktop/FLA/excel/FLA Return existing skeletal.xlsx",
        "/Users/apple/Desktop/FLA/skel_details.txt"
    )
