import openpyxl

wb = openpyxl.load_workbook("/Users/apple/Desktop/FLA/fla_automation_engine/fla_automation_engine/excel/FLA Return existing skeletal.xlsx", data_only=False)
sheet = wb["Section IV"]

print("=== Formulas in Section IV Row 39 (Equity Capital and PPS) ===")
print("Col D (PY):", sheet.cell(row=39, column=4).value)
print("Col E (FY):", sheet.cell(row=39, column=5).value)

print("=== Formulas in Section IV Row 30 (Net Worth of DIE) ===")
print("Col D (PY):", sheet.cell(row=30, column=4).value)
print("Col E (FY):", sheet.cell(row=30, column=5).value)

print("=== Formulas in Section IV Row 42 (Other Capital) ===")
print("Col D (PY):", sheet.cell(row=42, column=4).value)
print("Col E (FY):", sheet.cell(row=42, column=5).value)
