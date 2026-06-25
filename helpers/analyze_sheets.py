import openpyxl
import json

def analyze_skeletal(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    results = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        results[sheet_name] = []
        for r in range(1, sheet.max_row + 1):
            row_vals = []
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                row_vals.append({
                    "col": c,
                    "val": cell.value,
                    "formula": str(cell.value) if (isinstance(cell.value, str) and cell.value.startswith('=')) else None
                })
            # Check if there is any text in this row
            if any(cell["val"] is not None for cell in row_vals):
                results[sheet_name].append({
                    "row": r,
                    "cells": row_vals
                })
    wb.close()
    return results

def analyze_tool(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    results = {}
    for sheet_name in ['2_FINANCIALS', '3_FLA_RETURN']:
        sheet = wb[sheet_name]
        results[sheet_name] = []
        for r in range(1, sheet.max_row + 1):
            row_vals = []
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                row_vals.append({
                    "col": c,
                    "val": cell.value,
                    "formula": str(cell.value) if (isinstance(cell.value, str) and cell.value.startswith('=')) else None
                })
            if any(cell["val"] is not None for cell in row_vals):
                results[sheet_name].append({
                    "row": r,
                    "cells": row_vals
                })
    wb.close()
    return results

if __name__ == "__main__":
    skel = analyze_skeletal("/Users/apple/Desktop/FLA/automation_engine/automation_engine/excel/FLA Return existing skeletal.xlsx")
    tool = analyze_tool("/Users/apple/Desktop/FLA/automation_engine/excel/FLA_Tool_v5_fixed.xlsx")
    
    with open("/Users/apple/Desktop/FLA/skel_structure.json", "w") as f:
        json.dump(skel, f, indent=2)
    with open("/Users/apple/Desktop/FLA/tool_structure.json", "w") as f:
        json.dump(tool, f, indent=2)
        
    print("Skeletal and Tool structures analyzed and dumped to JSON.")
