import openpyxl
import pandas as pd

def inspect_formulas(file_path, sheet_name, num_rows=100):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb[sheet_name]
    
    print(f"\n=== Formulas / Values in {sheet_name} ===")
    rows_data = []
    for r in range(1, min(sheet.max_row + 1, num_rows + 1)):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
        # check if any element has a formula (starts with '=')
        has_formula = any(isinstance(val, str) and val.startswith('=') for val in row_vals)
        if any(row_vals):
            rows_data.append((r, row_vals, has_formula))
            
    for r, vals, has_form in rows_data[:50]:
        print(f"Row {r:03d} (has_formula={has_form}): {vals}")

if __name__ == "__main__":
    file_path = "/Users/apple/Desktop/FLA/excel/FLA_Tool_v5_fixed.xlsx"
    inspect_formulas(file_path, "2_FINANCIALS", 80)
    inspect_formulas(file_path, "3_FLA_RETURN", 80)
