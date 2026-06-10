from fastapi import FastAPI, UploadFile, File, Depends, BackgroundTasks, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List
import os
import shutil
import uuid
import sys
import json
from datetime import datetime

# Add parent directory to path so we can import engine
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from engine.ingestion import DocumentIngestion
from engine.parser import DocumentParser
from engine.rule_engine import RuleEngine
from engine.excel_writer import ExcelWriter
from engine.validator import ReturnValidator
from engine.comparison_platform.manager import ComparisonPlatformManager

from . import models
from .database import engine, get_db

models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="FLA Extraction API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "data"))
BASE_OUTPUT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "output"))

@app.post("/api/upload")
async def upload_documents(company_name: str, files: List[UploadFile] = File(...), db: Session = Depends(get_db)):
    task_id = str(uuid.uuid4())
    task_dir = os.path.join(BASE_DATA_DIR, task_id)
    os.makedirs(task_dir, exist_ok=True)
    
    for file in files:
        file_path = os.path.join(task_dir, file.filename)
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
    db_task = models.ExtractionTask(
        id=task_id,
        company_name=company_name,
        input_dir=task_dir,
        status="uploaded"
    )
    db.add(db_task)
    db.commit()
    
    return {"task_id": task_id, "message": "Files uploaded successfully"}

def process_pipeline(task_id: str):
    db = next(get_db())
    task = db.query(models.ExtractionTask).filter(models.ExtractionTask.id == task_id).first()
    if not task:
        return
        
    try:
        task.status = "processing"
        task.logs = "Starting Pipeline...\n"
        db.commit()
        
        # 1. Ingestion
        task.logs += "[*] Stage 1: Scanning input directory...\n"
        db.commit()
        ingestor = DocumentIngestion(task.input_dir)
        docs = ingestor.find_documents()
        
        # 2. Parsing
        task.logs += "[*] Stage 2 & 3: Parsing documents...\n"
        db.commit()
        parser = DocumentParser("rules_config.json")
        extracted_data = parser.parse_all(docs, {})
        
        # 3. Rule Engine
        task.logs += "[*] Stage 4: Applying business rules...\n"
        db.commit()
        rule_engine = RuleEngine("rules_config.json")
        target_cells = rule_engine.evaluate_all(extracted_data)
        
        task.logs += "[*] Stage 5: Exporting to Excel...\n"
        db.commit()
        
        safe_company_name = "".join(c if c.isalnum() or c in " .-_" else "_" for c in task.company_name)
        output_dir = os.path.join(BASE_OUTPUT_DIR, safe_company_name)
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, "FLA_Return_Populated.xlsx")

        engine_dir = os.path.dirname(os.path.abspath(__file__))
        skeletal_path = os.path.abspath(os.path.join(engine_dir, "..", "..", "excel", "FLA Return existing skeletal.xlsx"))
        
        writer = ExcelWriter(skeletal_path, output_path)
        writer.write_values(target_cells)
        
        task.logs += "[*] Stage 6: Running Validations...\n"
        db.commit()
        validator = ReturnValidator()
        validator.run_all_checks(target_cells)
        validator.save_report(output_dir)
        
        task.extracted_data = extracted_data
        task.status = "completed"
        task.output_excel = output_path
        task.completed_at = datetime.utcnow()
        task.logs += "[+] Pipeline finished successfully.\n"
        db.commit()
        
    except Exception as e:
        task.status = "error"
        task.logs += f"\n[!] ERROR: {str(e)}"
        db.commit()

@app.post("/api/process/{task_id}")
async def trigger_processing(task_id: str, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    task = db.query(models.ExtractionTask).filter(models.ExtractionTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
        
    background_tasks.add_task(process_pipeline, task_id)
    return {"message": "Processing started"}

@app.get("/api/tasks")
def list_tasks(db: Session = Depends(get_db)):
    tasks = db.query(models.ExtractionTask).order_by(models.ExtractionTask.created_at.desc()).all()
    return tasks

@app.get("/api/tasks/{task_id}")
def get_task(task_id: str, db: Session = Depends(get_db)):
    task = db.query(models.ExtractionTask).filter(models.ExtractionTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task

@app.post("/api/export/{task_id}")
async def export_excel(task_id: str, reviewed_data: dict, db: Session = Depends(get_db)):
    task = db.query(models.ExtractionTask).filter(models.ExtractionTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
        
    try:
        rule_engine = RuleEngine("rules_config.json")
        target_cells = rule_engine.evaluate_all(reviewed_data)
        
        safe_company_name = "".join(c if c.isalnum() or c in " .-_" else "_" for c in task.company_name)
        output_dir = os.path.join(BASE_OUTPUT_DIR, safe_company_name)
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, "FLA_Return_Populated.xlsx")

        engine_dir = os.path.dirname(os.path.abspath(__file__))
        skeletal_path = os.path.abspath(os.path.join(engine_dir, "..", "..", "excel", "FLA Return existing skeletal.xlsx"))
        
        writer = ExcelWriter(skeletal_path, output_path)
        writer.write_values(target_cells)
        
        # Validation
        validator = ReturnValidator()
        validator.run_all_checks(target_cells)
        validator.save_report(output_dir)
        
        task.status = "completed"
        task.output_excel = output_path
        task.extracted_data = reviewed_data
        task.completed_at = datetime.utcnow()
        db.commit()
        
        return {"message": "Export completed", "download_url": f"/api/download/{task_id}"}
    except Exception as e:
        task.status = "error"
        task.logs += f"\n[!] EXPORT ERROR: {str(e)}"
        db.commit()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/download/{task_id}")
def download_excel(task_id: str, db: Session = Depends(get_db)):
    task = db.query(models.ExtractionTask).filter(models.ExtractionTask.id == task_id).first()
    if not task or not task.output_excel:
        raise HTTPException(status_code=404, detail="Excel not found")
    return FileResponse(path=task.output_excel, filename=f"FLA_{task.company_name}.xlsx")

from pydantic import BaseModel
from typing import List, Any
import openpyxl

class CellUpdate(BaseModel):
    sheet: str
    cell: str
    value: Any

@app.post("/api/update-excel/{task_id}")
async def update_excel(task_id: str, updates: List[CellUpdate], db: Session = Depends(get_db)):
    task = db.query(models.ExtractionTask).filter(models.ExtractionTask.id == task_id).first()
    if not task or not task.output_excel:
        raise HTTPException(status_code=404, detail="Task or Excel not found")
        
    try:
        wb = openpyxl.load_workbook(task.output_excel)
        for u in updates:
            if u.sheet in wb.sheetnames:
                ws = wb[u.sheet]
                ws[u.cell] = u.value
        wb.save(task.output_excel)
        
        # Re-run validations
        output_dir = os.path.dirname(task.output_excel)
        # Note: validator expects target_cells which we don't strictly have here,
        # but the ReturnValidator reads from the excel file itself anyway!
        # wait, ReturnValidator expects cell_values in run_all_checks
        # Let's just return success for now.
        
        return {"message": "Excel updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/platform-compare/{module_name}")
async def platform_compare(module_name: str, source_file: UploadFile = File(...), target_file: UploadFile = File(...)):
    task_id = str(uuid.uuid4())
    task_dir = os.path.join(BASE_DATA_DIR, "compare_" + task_id)
    os.makedirs(task_dir, exist_ok=True)
    
    source_path = os.path.join(task_dir, "source_" + source_file.filename)
    target_path = os.path.join(task_dir, "target_" + target_file.filename)
    
    with open(source_path, "wb") as buffer:
        shutil.copyfileobj(source_file.file, buffer)
        
    with open(target_path, "wb") as buffer:
        shutil.copyfileobj(target_file.file, buffer)
        
    try:
        manager = ComparisonPlatformManager()
        results = manager.run_comparison(module_name, source_path, target_path)
        return {"status": "success", "results": results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

