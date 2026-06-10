import openpyxl
from openpyxl.utils import get_column_letter
import os

class ExcelWriter:
    def __init__(self, skeletal_path, output_path=None):
        self.skeletal_path = skeletal_path
        if output_path is None:
            # Save in the same folder as skeletal but with a new name
            dir_name = os.path.dirname(skeletal_path)
            self.output_path = os.path.join(dir_name, "FLA Return Populated.xlsx")
        else:
            self.output_path = output_path
            
    def write_values(self, cell_values):
        """Copies skeletal Excel to output_path, writes calculated/extracted values on the copy, and saves it."""
        import shutil
        
        if not os.path.exists(self.skeletal_path):
            raise FileNotFoundError(f"Skeletal Excel not found at {self.skeletal_path}")
            
        # Ensure output directory exists
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        
        # Copy skeletal template to the output path first
        print(f"[*] Copying skeletal template to: {self.output_path}")
        shutil.copy(self.skeletal_path, self.output_path)
        
        print(f"[*] Loading copied target Excel: {self.output_path}")
        wb = openpyxl.load_workbook(self.output_path, data_only=False)
        
        # Iterate over sections
        # Iterate over sections
        for section, cells in cell_values.items():
            if section not in wb.sheetnames:
                print(f"[!] Warning: Sheet '{section}' not found in skeletal Excel. Skipping...")
                continue
                
            sheet = wb[section]
            print(f"[*] Populating sheet '{section}' with {len(cells)} active fields...")
            
            # Specific coordinates that are merged in the skeletal template but need independent PY/FY values
            unmerge_targets = {
                "Section II": ["C5:G5", "C6:G6", "C11:G11", "C24:G24", "C30:G30", "C34:G34"],
                "Section III": ["C20:E20", "C23:E23", "C44:E44", "C47:E47"],
                "Section IV": ["C39:E39", "C42:E42"]
            }
            
            # Custom dynamic row insertion for Section III multiple DI countries
            num_new_rows = 0
            if section == "Section III" and "fdi_investor_2_countries_json" in cells:
                import json
                import re
                try:
                    countries_data = json.loads(cells["fdi_investor_2_countries_json"])
                    if len(countries_data) > 1:
                        num_new_rows = len(countries_data) - 1
                        
                        # Insert rows after row 41
                        print(f"[*] Section III: Inserting {num_new_rows} new rows for multi-country DI consolidation...")
                        sheet.insert_rows(42, num_new_rows)
                        
                        # Manually shift all merged ranges below row 41 by num_new_rows due to openpyxl insert_rows bug
                        merged_ranges = list(sheet.merged_cells.ranges)
                        sheet.merged_cells.ranges = []
                        for r in merged_ranges:
                            if r.min_row > 41:
                                r.shift(row_shift=num_new_rows)
                            sheet.merged_cells.add(r)
                        
                        # Copy styles/borders from Row 41 to newly inserted rows
                        from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
                        for i in range(1, num_new_rows + 1):
                            target_row = 41 + i
                            sheet.row_dimensions[target_row].height = 24
                            for col_idx in range(1, sheet.max_column + 1):
                                src_cell = sheet.cell(row=41, column=col_idx)
                                dst_cell = sheet.cell(row=target_row, column=col_idx)
                                # Copy font, border, fill, alignment, number format
                                if src_cell.has_style:
                                    dst_cell.font = Font(name=src_cell.font.name, size=src_cell.font.size, bold=src_cell.font.bold, italic=src_cell.font.italic, color=src_cell.font.color)
                                    dst_cell.border = Border(left=src_cell.border.left, right=src_cell.border.right, top=src_cell.border.top, bottom=src_cell.border.bottom)
                                    dst_cell.fill = PatternFill(fill_type=src_cell.fill.fill_type, start_color=src_cell.fill.start_color, end_color=src_cell.fill.end_color)
                                    dst_cell.alignment = Alignment(horizontal=src_cell.alignment.horizontal, vertical=src_cell.alignment.vertical, wrap_text=src_cell.alignment.wrap_text)
                                    dst_cell.number_format = src_cell.number_format
                        
                        # Populate country/percentages in Row 41 and the newly inserted rows
                        for idx, country_info in enumerate(countries_data):
                            row_num = 41 + idx
                            c_name = country_info["country"]
                            p_py = float(country_info.get("percent_py", 0))
                            p_fy = float(country_info.get("percent_fy", 0))
                            
                            sheet.cell(row=row_num, column=2, value=c_name) # Column B
                            
                            cell_py = sheet.cell(row=row_num, column=3, value=p_py / 100.0) # Column C
                            cell_py.number_format = '0.00%'
                            
                            cell_fy = sheet.cell(row=row_num, column=4, value=p_fy / 100.0) # Column D
                            cell_fy.number_format = '0.00%'
                            
                        # Shift all coordinates > 41 in cells by num_new_rows!
                        new_cells = {}
                        for coord, val in cells.items():
                            if coord == "fdi_investor_2_countries_json":
                                continue
                            # Parse coord to get row
                            match = re.match(r"^([a-zA-Z]+)(\d+)$", coord)
                            if match:
                                col_part, row_part = match.groups()
                                row_num = int(row_part)
                                if row_num == 41:
                                    continue
                                elif row_num > 41:
                                    new_coord = f"{col_part}{row_num + num_new_rows}"
                                    new_cells[new_coord] = val
                                else:
                                    new_cells[coord] = val
                            else:
                                new_cells[coord] = val
                        cells = new_cells
                        
                        # Dynamically shift unmerge_targets coordinates for Section III
                        sec3_unmerge = []
                        for rng_str in unmerge_targets["Section III"]:
                            parts = rng_str.split(":")
                            new_parts = []
                            for p in parts:
                                match = re.match(r"^([a-zA-Z]+)(\d+)$", p)
                                if match:
                                    col_part, row_part = match.groups()
                                    row_num = int(row_part)
                                    if row_num > 41:
                                        new_parts.append(f"{col_part}{row_num + num_new_rows}")
                                    else:
                                        new_parts.append(p)
                                else:
                                    new_parts.append(p)
                            sec3_unmerge.append(":".join(new_parts))
                        unmerge_targets["Section III"] = sec3_unmerge
                except Exception as e:
                    print(f"[!] Error processing multi-country DI insertion: {e}")
                
                # ALWAYS remove the json key so openpyxl doesn't try to write to it
                if "fdi_investor_2_countries_json" in cells:
                    del cells["fdi_investor_2_countries_json"]
            
            if section in unmerge_targets:
                for rng_str in unmerge_targets[section]:
                    try:
                        # Check if it is currently merged in the sheet
                        for rng in list(sheet.merged_cells.ranges):
                            if rng.coord == rng_str:
                                sheet.unmerge_cells(rng_str)
                                # Set the top-left cell explicitly
                                top_left = rng_str.split(":")[0]
                                sheet[top_left] = "Auto-calculated"
                                break
                    except Exception as e:
                        print(f"[!] Warning: Could not unmerge {rng_str} in sheet {section}: {e}")
            
            # Map merged cells to their top-left parent cells
            merged_map = {}
            for rng in sheet.merged_cells.ranges:
                top_left = rng.start_cell.coordinate
                for r in range(rng.min_row, rng.max_row + 1):
                    for c in range(rng.min_col, rng.max_col + 1):
                        cell_coord = get_column_letter(c) + str(r)
                        merged_map[cell_coord] = top_left
            
            for coord, val in cells.items():
                resolved_coord = merged_map.get(coord, coord)
                try:
                    # Preserve template structure: do not overwrite cells containing "Auto-calculated"
                    existing_val = sheet[resolved_coord].value
                    if isinstance(existing_val, str) and "auto-calculated" in existing_val.lower():
                        # Skipping to retain the merged cell structure and visual "Auto-calculated" label
                        continue
                        
                    # If this is a percentage coordinate, convert to decimal and apply percentage formatting
                    if (section == "Section II" and resolved_coord in ["F24", "G24"]) or \
                       (section == "Section III" and resolved_coord in ["D17", "E17", "C41", "D41"]) or \
                       (section == "Section IV" and resolved_coord in ["E19", "F19"]):
                        try:
                            val_float = float(val)
                            val = val_float / 100.0
                            sheet[resolved_coord].number_format = '0.00%'
                        except Exception:
                            pass
                            
                    # Write to the resolved coordinate (always the top-left of the merged block)
                    sheet[resolved_coord] = val
                except Exception as e:
                    print(f"[!] Error writing {val} to cell {coord} (resolved: {resolved_coord}) in sheet {section}: {e}")
                    
        # Apply premium formatting fixes to prevent clipping, overlapping, or jagged borders
        self.beautify_layout(wb)
                    
        # Save output
        print(f"[+] Saving populated Excel to: {self.output_path}")
        wb.save(self.output_path)
        return self.output_path

    def beautify_layout(self, wb):
        """Applies premium formatting fixes to prevent clipping, overlapping, or jagged borders."""
        from openpyxl.styles import Border, Side, Alignment
        
        # 1. Section III Beautifications
        if "Section III" in wb.sheetnames:
            sheet = wb["Section III"]
            # Row 16 contains a very long description in C16. Increase row height to avoid clipping!
            sheet.row_dimensions[16].height = 48
            # Row 17 (data row) can have a normal, clean height
            sheet.row_dimensions[17].height = 24
            
            # Align filled cells beautifully
            for r in [17, 21, 22, 24, 25, 26]:
                for c in [2, 3, 4, 5, 6, 7]:
                    cell = sheet.cell(row=r, column=c)
                    # Align: left/center for text, center for country/numbers
                    if c in [2, 3]:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    elif c in [4, 5]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Unify table borders up to column G (columns 1 to 7) for rows 15 to 26
            # to remove the jaggedness on the right side of the sheet!
            for r in range(15, 27):
                for c in range(1, 8):
                    cell = sheet.cell(row=r, column=c)
                    current_border = cell.border
                    new_left = current_border.left if current_border.left.style else Side(style='thin', color='000000')
                    new_right = current_border.right if current_border.right.style else Side(style='thin', color='000000')
                    new_top = current_border.top if current_border.top.style else Side(style='thin', color='000000')
                    new_bottom = current_border.bottom if current_border.bottom.style else Side(style='thin', color='000000')
                    
                    cell.border = Border(left=new_left, right=new_right, top=new_top, bottom=new_bottom)
                    
        # 2. Section IV Beautifications
        if "Section IV" in wb.sheetnames:
            sheet = wb["Section IV"]
            # Set DIE 1 details row heights and alignment
            sheet.row_dimensions[19].height = 24
            
            # Align filled cells beautifully
            for r in [19, 23, 24, 26, 27, 28]:
                for c in [1, 2, 3, 4, 5, 6]:
                    cell = sheet.cell(row=r, column=c)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
            # Unify borders up to column G for rows 17 to 29
            for r in range(17, 30):
                for c in range(1, 8):
                    cell = sheet.cell(row=r, column=c)
                    current_border = cell.border
                    new_left = current_border.left if current_border.left.style else Side(style='thin', color='000000')
                    new_right = current_border.right if current_border.right.style else Side(style='thin', color='000000')
                    new_top = current_border.top if current_border.top.style else Side(style='thin', color='000000')
                    new_bottom = current_border.bottom if current_border.bottom.style else Side(style='thin', color='000000')
                    
                    cell.border = Border(left=new_left, right=new_right, top=new_top, bottom=new_bottom)
