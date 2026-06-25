import openpyxl

wb = openpyxl.load_workbook("/Users/apple/Desktop/FLA/automation_engine/automation_engine/excel/FLA Return existing skeletal.xlsx", data_only=True)
sheet = wb["Section III"]

print("=== SKELETAL SPREADSHEET ROWS 20-25 COLUMNS F & G ===")
for r in range(20, 26):
    row_str = []
    for c in [6, 7]:
        col_letter = openpyxl.utils.get_column_letter(c)
        cell = sheet.cell(row=r, column=c)
        row_str.append(f"{col_letter}{r}: Value='{cell.value}'")
    print(f"Row {r:02d}: " + " | ".join(row_str))
